#!/usr/bin/env python3
"""
Unhide the 'Matrix Table' worksheet in a Hudson input Excel file and read it into a pandas DataFrame.

Usage:
  python scripts/unhide_matrix_table.py -i "path/to/hudson_input.xlsx|.xlsm" [-o "path/to/output.xlsx|.xlsm"]

If -o/--output is not provided, the input file will be modified in place.
The script prints a small preview of the DataFrame and its shape.
"""
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


DEFAULT_SHEET_NAME = "Matrix Table"


def unhide_sheet(xlsx_path: Path, sheet_name: str = DEFAULT_SHEET_NAME, output_path: Path | None = None) -> Path:
    """Unhide a worksheet in an Excel file and save the workbook.

    Args:
        xlsx_path: Path to the .xlsx file.
        sheet_name: Name of the worksheet to unhide.
        output_path: Optional path to save the modified workbook. If None, saves in place.

    Returns:
        Path to the saved workbook.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    wb = load_workbook(filename=xlsx_path, read_only=False, keep_links=True, keep_vba=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Worksheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    # Set sheet state to visible regardless of current state (visible, hidden, veryHidden)
    try:
        ws.sheet_state = "visible"
    except Exception as e:
        # Provide a clearer message if workbook/worksheet is protected
        raise RuntimeError(
            "Failed to change sheet visibility. The workbook or sheet may be protected."
        ) from e

    save_path = Path(output_path) if output_path else xlsx_path
    # Ensure parent directory exists when saving to a new file path
    save_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(save_path)
    return save_path


essential_read_excel_kwargs = dict(engine="openpyxl")


def read_matrix_table(xlsx_path: Path, sheet_name: str = DEFAULT_SHEET_NAME) -> pd.DataFrame:
    """Read the specified worksheet into a pandas DataFrame.

    Args:
        xlsx_path: Path to the workbook containing the sheet.
        sheet_name: Name of the worksheet to read.

    Returns:
        pandas.DataFrame with the sheet contents.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, **essential_read_excel_kwargs)
    return df


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Unhide the 'Matrix Table' sheet and read it into a pandas DataFrame."
    )
    parser.add_argument(
        "-i",
        "--input",
        required=True,
        help="Path to the Hudson input Excel .xlsx/.xlsm file",
    )
    parser.add_argument(
        "-o",
        "--output",
        required=False,
        default=None,
        help=(
            "Optional path to save a copy of the workbook with the sheet visible. "
            "If omitted, modifies the input file in place."
        ),
    )
    parser.add_argument(
        "-s",
        "--sheet-name",
        required=False,
        default=DEFAULT_SHEET_NAME,
        help="Worksheet name to unhide and read (default: 'Matrix Table')",
    )
    parser.add_argument(
        "--preview-rows",
        type=int,
        default=5,
        help="Number of rows to preview from the DataFrame (default: 5)",
    )

    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output) if args.output else None

    # 1) Unhide the sheet and save
    saved_path = unhide_sheet(input_path, sheet_name=args.sheet_name, output_path=output_path)

    # 2) Read into DataFrame
    df = read_matrix_table(saved_path, sheet_name=args.sheet_name)

    # 3) Print a small preview and shape
    print(f"Read DataFrame from '{args.sheet_name}' in {saved_path}")
    with pd.option_context("display.max_columns", 20, "display.width", 120):
        print(df.head(args.preview_rows))
    print(f"DataFrame shape: {df.shape}")


if __name__ == "__main__":
    main()

