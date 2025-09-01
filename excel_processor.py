import sys
import os
from pathlib import Path
import re
from datetime import datetime, date
import shutil
import openpyxl
import requests

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils.datetime import from_excel as excel_from_serial
    from openpyxl.utils import get_column_letter
    from dotenv import load_dotenv
    from graph_auth import acquire_graph_token
except ImportError as e:
    print('DEPENDENCY_ERROR: Required libraries are not installed.')
    print('Please install: pip install pandas openpyxl python-dotenv requests msal')
    print(e)
    sys.exit(1)

# Load environment variables
load_dotenv()

# --- Define File Paths and Constants ---
# These paths point to the files as they were referenced in your original scripts.
# Update them if the file names or locations change.
SRC_UNFILTERED_NAME = '2-copy-reformat/ERCOT-new.xlsx'
DST_MASTER_TABLE_NAME = '2-copy-reformat/Master-Table.xlsx'

TARGET_TERMS = {12, 24, 36, 48, 60}

# Columns that will be written to the master table after filtering
BASE_COLS = [
    'Start Month',
    'State',
    'Utility',
    'Congestion Zone',
    'Load Factor',
    'Term',
    'Product',
    '0-200,000',
]

# --- Backup Functions ---

def create_master_table_backup(master_path):
    """Create a timestamped backup copy of the master table before modifications.

    Args:
        master_path (Path): Path to the master table file

    Returns:
        Path: Path to the backup file if successful, None if failed
    """
    if not master_path.exists():
        print(f"Master table not found for backup: {master_path}")
        return None

    try:
        # Create backup directory if it doesn't exist
        backup_dir = master_path.parent / "backups"
        backup_dir.mkdir(exist_ok=True)

        # Generate timestamped backup filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{master_path.stem}_backup_{timestamp}{master_path.suffix}"
        backup_path = backup_dir / backup_name

        # Copy the master table to backup location
        shutil.copy2(master_path, backup_path)
        print(f"Created backup: {backup_path}")
        return backup_path

    except Exception as e:
        print(f"Failed to create backup of master table: {e}")
        return None


# --- SharePoint Integration Functions ---

def download_sharepoint_file(file_name, download_path=None):
    """Download a file from SharePoint using Microsoft Graph API

    Args:
        file_name (str): Name of the file to download from SharePoint
        download_path (str, optional): Local path where to save the file.
                                     If None, saves to current directory with same name.

    Returns:
        Path: Path to the downloaded file if successful, None if failed
    """
    # Get credentials from environment
    tenant_id = os.getenv("TENANT_ID")
    client_id = os.getenv("CLIENT_ID")
    client_secret = os.getenv("CLIENT_SECRET")
    site_hostname = os.getenv("SITE_HOSTNAME")
    site_path = os.getenv("SITE_PATH")
    sharepoint_folder = os.getenv("SHAREPOINT_UPLOAD_FOLDER")

    if not all([tenant_id, client_id, client_secret, site_hostname, site_path, sharepoint_folder]):
        print("ERROR: Missing SharePoint configuration in environment variables.")
        print("Required: TENANT_ID, CLIENT_ID, CLIENT_SECRET, SITE_HOSTNAME, SITE_PATH, SHAREPOINT_UPLOAD_FOLDER")
        return None

    # Determine download path
    if download_path is None:
        download_path = Path(file_name)
    else:
        download_path = Path(download_path)

    try:
        # Acquire access token
        token = acquire_graph_token(tenant_id, client_id, client_secret)
        headers = {"Authorization": f"Bearer {token['access_token']}"}

        # First, get the site ID
        site_url = f"https://graph.microsoft.com/v1.0/sites/{site_hostname}:/sites/{site_path}"
        site_response = requests.get(site_url, headers=headers)

        if site_response.status_code != 200:
            print(f"Failed to get site info. Status: {site_response.status_code}")
            print(f"Response: {site_response.text}")
            return None

        site_data = site_response.json()
        site_id = site_data['id']
        print(f"Found SharePoint site ID: {site_id}")

        # Get the default drive (document library)
        drive_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive"
        drive_response = requests.get(drive_url, headers=headers)

        if drive_response.status_code != 200:
            print(f"Failed to get drive info. Status: {drive_response.status_code}")
            print(f"Response: {drive_response.text}")
            return None

        drive_data = drive_response.json()
        drive_id = drive_data['id']
        print(f"Found drive ID: {drive_id}")

        # Construct the file path and download URL
        file_path = f"{sharepoint_folder}/{file_name}".replace('//', '/')
        if file_path.startswith('/'):
            file_path = file_path[1:]  # Remove leading slash

        file_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{file_path}:/content"
        print(f"Downloading from SharePoint: {file_url}")

        # Download the file
        response = requests.get(file_url, headers=headers)

        if response.status_code == 200:
            with open(download_path, "wb") as local_file:
                local_file.write(response.content)
            print(f"Successfully downloaded from SharePoint: {download_path}")
            return download_path
        else:
            print(f"Failed to download file from SharePoint. Status: {response.status_code}")
            print(f"Response: {response.text}")
            return None

    except Exception as e:
        print(f"Error during SharePoint download: {str(e)}")
        return None


def create_master_table_if_not_exists(master_path):
    """Create a basic master table with headers if it doesn't exist"""
    if master_path.exists():
        return True

    try:
        # Create directory if it doesn't exist
        master_path.parent.mkdir(parents=True, exist_ok=True)

        # Create a new workbook with headers
        wb = openpyxl.Workbook()
        ws = wb.active

        # Add headers for BASE_COLS (columns B-I) with ID in column A
        headers = ['ID'] + BASE_COLS
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header)

        wb.save(master_path)
        print(f"Created new master table: {master_path}")
        return True
    except Exception as e:
        print(f"Failed to create master table: {e}")
        return False


def download_master_table_from_sharepoint(dst_master_path=None):
    """Download the master table from SharePoint if it doesn't exist locally

    Args:
        dst_master_path (Path, optional): Path where to save the master table

    Returns:
        bool: True if master table is available (downloaded or already exists), False otherwise
    """
    if dst_master_path is None:
        dst_master_path = Path(DST_MASTER_TABLE_NAME)

    # If master table already exists locally, no need to download
    if dst_master_path.exists():
        print(f"Master table already exists: {dst_master_path}")
        return True

    # excel_files Save existing MPN_file before making a new file
    # ap1_data = pd.read_excel(DST_MASTER_TABLE_NAME)
    # d = "MPN_file_"+datetime.now().strftime("%b_%d_%Y-%H-%M-%S")
    # dirr = 'C:/Users/TThurston/Desktop/2023/Automation_testing/New_Files/' + d +'.xlsx'
    # ap1_data.to_excel(dirr, index=False)

    # Try to download the master table from SharePoint
    master_file_name = os.getenv("DAILY_PRICING_FILE_NAME", "DAILY PRICING - new.xlsx")
    print(f"Master table not found locally. Attempting to download from SharePoint: {master_file_name}")

    # Create the directory if it doesn't exist
    dst_master_path.parent.mkdir(parents=True, exist_ok=True)

    # Download master table from the parent directory (not Input Files subfolder)
    # Temporarily override the sharepoint folder for master table download
    original_folder = os.getenv("SHAREPOINT_UPLOAD_FOLDER")
    master_folder = "/Kilowatt/Client Pricing Sheets"  # Parent directory

    # Temporarily set environment variable for master table download
    os.environ["SHAREPOINT_UPLOAD_FOLDER"] = master_folder
    try:
        downloaded_file = download_sharepoint_file(master_file_name, str(dst_master_path))
    finally:
        # Restore original folder setting
        if original_folder:
            os.environ["SHAREPOINT_UPLOAD_FOLDER"] = original_folder
    if downloaded_file:
        print(f"Successfully downloaded master table from SharePoint to: {dst_master_path}")
        return True
    else:
        print(f"Failed to download master table from SharePoint: {master_file_name}")
        return False


def download_and_process_sharepoint_file(file_name, dst_master_path=None, sheet_name_prefer=None):
    """Download a file from SharePoint and process it directly

    Args:
        file_name (str): Name of the file to download from SharePoint
        dst_master_path (Path, optional): Path to master table. Defaults to DST_MASTER_TABLE_NAME
        sheet_name_prefer (str, optional): Preferred sheet name for processing

    Returns:
        int: Number of rows appended to master table
    """
    if dst_master_path is None:
        dst_master_path = Path(DST_MASTER_TABLE_NAME)

    # Ensure master table is available (download from SharePoint if needed)
    if not download_master_table_from_sharepoint(dst_master_path):
        print("Cannot proceed without master table")
        return 0

    # Download the file from SharePoint
    downloaded_file = download_sharepoint_file(file_name)
    if downloaded_file is None:
        print(f" line 234  Failed to download {file_name} from SharePoint")
        return 0

    try:
        # Process the downloaded file
        print(downloaded_file)
        if downloaded_file.suffix.lower() == '.xlsm':
            rows_appended = process_xlsm_file(downloaded_file, dst_master_path, sheet_name_prefer)
        else:
            # For other Excel files, use the standard filtering approach
            try:
                sheets = pd.read_excel(downloaded_file, sheet_name=None)
                filtered_data = []
                for _, df in sheets.items():
                    filtered_data.append(filter_sheet(df))

                combined_df = pd.concat(filtered_data, ignore_index=True)
                rows_appended = append_filtered_dataframe_to_master(combined_df, dst_master_path)
            except Exception as e:
                print(f'Error processing downloaded file: {e}')
                return 0

        print(f"Successfully processed {downloaded_file.name}: {rows_appended} rows appended")
        return rows_appended

    finally:
        # Optionally clean up the downloaded file
        # Uncomment the next line if you want to delete the file after processing
        # downloaded_file.unlink(missing_ok=True)
        pass


# --- Core Functions from Original Scripts ---

def parse_term_to_int(val):
    if pd.isna(val) or val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val) if float(int(val)) == val else None
    s = str(val).strip()
    if s.isdigit():
        return int(s)
    m = re.search(r"(\d+)", s)
    return int(m.group(1)) if m else None

def last_data_row(ws, scan_cols=40):
    max_r = ws.max_row
    for r in range(max_r, 0, -1):
        for c in range(1, scan_cols + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ''):
                return r
    return 0

def find_first_blank_row(ws):
    """Find the first blank row in column A (ID column) to start writing data."""
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v in (None, ''):
            return r
    # If no blank row found, return the next row after the last row
    return ws.max_row + 1

def get_next_id(ws):
    """Get the next ID number by finding the maximum ID in column A and adding 1."""
    max_id = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, (int, float)):
            try:
                vi = int(v)
            except Exception:
                continue
            if vi > max_id:
                max_id = vi
    return max_id + 1

def filter_sheet(df):
    prod_col = next((col for col in df.columns if col.strip().lower() in ['product', 'products']), None)
    term_col = next((col for col in df.columns if col.strip().lower() in ['term', 'terms']), None)

    if prod_col is None or term_col is None:
        return pd.DataFrame(columns=BASE_COLS)

    prod_mask = df[prod_col].astype(str).str.strip().str.lower().eq('fixed price')
    term_vals = df[term_col].map(parse_term_to_int)
    term_mask = term_vals.isin(TARGET_TERMS)

    cols = [c for c in BASE_COLS if c in df.columns]
    filtered = df.loc[prod_mask & term_mask, cols].copy()

    return filtered

def add_formulas(path, start_row, end_row):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        if end_row < 2:
            continue

        # Add headers if the workbook is new
        headers = {
            9: 'I (blank)', 10: 'J Index', 11: 'K Concat', 12: 'L ConstDate',
            13: 'M =B', 14: 'N Region', 15: 'O LF Norm', 16: 'P Supplier',
            17: 'Q TermMonths', 18: 'R (skip)', 19: 'S', 20: 'T',
            21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y', 26: 'Z', 27: 'AA',
        }
        if start_row == 1:
            for col_idx, name in headers.items():
                ws.cell(row=1, column=col_idx, value=name)

        for r in range(start_row, end_row + 1):
            # J: index (J2 = 1; Jn = J(n-1)+1)
            if r == 1:
                ws.cell(row=r, column=10, value=1)
            else:
                ws.cell(row=r, column=10, value=f"=J{r-1}+1")

            # ... [all other formulas from your build_ercot script] ...

            # V: T+U
            ws.cell(row=r, column=22, value=f"=T{r}+U{r}")
            # AA: 10
            ws.cell(row=r, column=27, value="=10")

    wb.save(path)


# --- Append L..AA from source to B..Q in destination, with A as sequence and O/P swap ---

def append_l_aa(src_path: Path, dst_path: Path, sheet_index: int = 0) -> None:
    """
    Read first sheet of src_path and take columns L..AA (12..27).
    Append to dst_path first sheet starting at next empty row:
      - Write into columns B..Q (2..17)
      - Column A is previous A + 1 (sequence)
      - Swap O and P from the source block (indices 3 and 4 within L..AA)
      - Convert Excel serial dates to real dates and set number_format to 'm/dd/yyyy'
    """
    # Validate files
    if not Path(src_path).exists():
        print(f"ERROR: Source not found: {src_path}")
        return
    if not Path(dst_path).exists():
        print(f"ERROR: Destination not found: {dst_path}")
        return

    # Load workbooks
    wb_src = load_workbook(src_path, data_only=True)
    ws_src = wb_src.worksheets[sheet_index]

    wb_dst = load_workbook(dst_path)
    ws_dst = wb_dst.worksheets[0]

    # Gather L..AA rows with any non-empty value
    src_rows: list[list] = []
    max_r = ws_src.max_row
    for r in range(2, max_r + 1):  # skip header
        row_vals = [ws_src.cell(row=r, column=c).value for c in range(12, 28)]
        if any(v not in (None, '') for v in row_vals):
            src_rows.append(row_vals)

    if not src_rows:
        print("No data to append from source (L..AA all blank).")
        return

    # Determine destination starting row
    last_row = last_data_row(ws_dst, scan_cols=50)
    start_row = last_row + 1 if last_row >= 1 else 1
    print(f"Destination last data row: {last_row}. Appending starting at row {start_row} in columns A..R")

    # Append rows
    for i, row_vals in enumerate(src_rows):
        r_dst = start_row + i

        # Column A sequence
        if r_dst == 1:
            ws_dst.cell(row=r_dst, column=1, value=1)
        else:
            prev = ws_dst.cell(row=r_dst - 1, column=1).value
            try:
                prev_int = int(prev) if prev is not None and str(prev).strip() != '' else 0
            except Exception:
                prev_int = 0
            ws_dst.cell(row=r_dst, column=1, value=prev_int + 1)

        # Swap O and P (indices 3 and 4 within L..AA)
        row_vals_to_write = list(row_vals)
        if len(row_vals_to_write) >= 5:
            row_vals_to_write[3], row_vals_to_write[4] = row_vals_to_write[4], row_vals_to_write[3]

        # Apply master formats to the entire row first
        apply_master_formats(ws_dst, r_dst)

        # Write into B..Q with date conversion/formatting
        for offset, val in enumerate(row_vals_to_write):
            is_date = False
            # Excel serial number -> datetime/date (force date-only)
            if isinstance(val, (int, float)) and 20000 < float(val) < 60000:
                try:
                    dt = excel_from_serial(float(val))
                    val = dt.date()
                    is_date = True
                except Exception:
                    pass
            # Already a Python datetime/date -> ensure date-only
            elif isinstance(val, datetime):
                val = val.date()
                is_date = True
            elif isinstance(val, date):
                is_date = True
            # Strings that look like dates/timestamps -> parse and coerce to date
            elif isinstance(val, str):
                s = str(val).strip()
                try:
                    dt = pd.to_datetime(s, errors='coerce')
                    if pd.notna(dt):
                        # Convert pandas Timestamp/Datetime to Python date
                        try:
                            dt_py = dt.to_pydatetime()
                        except AttributeError:
                            dt_py = dt
                        val = dt_py.date()
                        is_date = True
                except Exception:
                    pass

            cell = ws_dst.cell(row=r_dst, column=2 + offset, value=val)
            # Note: Master formats are already applied above, so we don't override them here
            # unless it's a special case like dates that need specific handling

    try:
        wb_dst.save(dst_path)
        print(f"Appended {len(src_rows)} rows (L..AA -> B..Q) to {dst_path.name}.")
    except PermissionError:
        print("ERROR: Could not save destination file. If it is open in Excel or locked, please close it and re-run.")
        return

# --- Main Logic ---

# --- Helpers to process .xlsm inputs and append using the same logic ---

def append_filtered_dataframe_to_master(combined_df: 'pd.DataFrame', dst: Path) -> int:
    """Append filtered rows in combined_df into the master table at dst.
    Uses proper column mapping based on master table structure.
    Returns number of rows appended.
    """
    if combined_df is None or combined_df.empty:
        print("No data to append after filtering.")
        return 0

    # Create backup before modifying master table
    backup_path = create_master_table_backup(dst)
    if backup_path is None:
        print("Warning: Could not create backup, proceeding anyway...")
    else:
        print(f"Master table backed up to: {backup_path}")

    # Master table column structure (B-Q columns)
    MASTER_HEADERS = [
        'Price_Date','Date','Zone','REP1','Load','Term','Min_MWh','Max_MWh',
        'Daily_No_Ruc','RUC_Nodal','Daily','Com_Disc','HOA_Disc','Broker_Fee','Meter_Fee','Max_Meters'
    ]



    wb_dst = load_workbook(dst)
    ws_dst = wb_dst.active

    # Determine current max ID in column A
    max_id = 0
    for r in range(2, ws_dst.max_row + 1):
        v = ws_dst.cell(row=r, column=1).value
        if isinstance(v, (int, float)):
            try:
                vi = int(v)
                if vi > max_id:
                    max_id = vi
            except Exception:
                continue
    next_id = max_id + 1

    # Find the first blank ID row (column A) to start writing
    first_blank_row = None
    for r in range(2, ws_dst.max_row + 1):
        v = ws_dst.cell(row=r, column=1).value
        if v in (None, ''):
            first_blank_row = r
            break
    if first_blank_row is None:
        first_blank_row = ws_dst.max_row + 1



    # Create mapping from BASE_COLS to MASTER_HEADERS positions
    # BASE_COLS: ['Start Month', 'State', 'Utility', 'Congestion Zone', 'Load Factor', 'Term', 'Product', '0-200,000']
    # MASTER_HEADERS: ['Price_Date','Date','Zone','REP1','Load','Term','Min_MWh','Max_MWh', ...]
    col_mapping = {
        'Start Month': 1,    # Date (column C)
        'State': None,       # No direct mapping
        'Utility': 4,        # REP1 (column F)
        'Congestion Zone': 2, # Zone (column D)
        'Load Factor': 3,    # Load (column E)
        'Term': 5,           # Term (column G)
        'Product': None,     # No direct mapping
        '0-200,000': None,   # No direct mapping
    }

    # Append filtered data rows
    rows_appended = 0
    write_row = first_blank_row

    for r_offset, row_data in enumerate(combined_df.itertuples(index=False), start=0):
        dst_row = write_row + r_offset

        # Column A: ID
        ws_dst.cell(row=dst_row, column=1, value=next_id + r_offset)

        # Apply formats to whole row A..Q first
        apply_master_formats(ws_dst, dst_row)

        # Map filtered data to master table columns
        for i, col_name in enumerate(BASE_COLS):
            if i < len(row_data):
                value = row_data[i]
                master_col_idx = col_mapping.get(col_name)
                if master_col_idx is not None:
                    # Write to the mapped column (add 1 for B=2, C=3, etc.)
                    ws_dst.cell(row=dst_row, column=master_col_idx + 2, value=value)

        rows_appended += 1

    # Save
    try:
        wb_dst.save(dst)
    except PermissionError:
        print("ERROR: Could not save destination file. Please close it if it's open and re-run.")
        sys.exit(4)

    return rows_appended


def a(master_df, dst_master_path):
    """Append a DataFrame that already has master table column structure to the master table.

    Args:
        master_df: DataFrame with columns matching master table structure (ID, Price_Date, Date, Zone, etc.)
        dst_master_path: Path to the master table Excel file

    Returns:
        int: Number of rows appended
    """
    from openpyxl import load_workbook

    if master_df is None or master_df.empty:
        print("No data to append after filtering.")
        return 0

    # Create backup before modifying master table
    backup_path = create_master_table_backup(dst_master_path)
    if backup_path is None:
        print("Warning: Could not create backup, proceeding anyway...")
    else:
        print(f"Master table backed up to: {backup_path}")

    # Master table column structure (B-Q columns, excluding ID in column A)
    MASTER_HEADERS = [
        'Price_Date','Date','Zone','Load','REP1','Term','Min_MWh','Max_MWh',
        'Daily_No_Ruc','RUC_Nodal','Daily','Com_Disc','HOA_Disc','Broker_Fee','Meter_Fee','Max_Meters'
    ]

    # Load the master workbook
    wb_dst = load_workbook(dst_master_path)
    ws_dst = wb_dst.active

    from datetime import date
    # Column B (Price_Date) should be today's date - this is already set correctly in the transformation
    # Column C (Date) should be the start date from input file - keep the original transformed value
    # Do NOT override the Date column here as it should contain the start date from input

    master_df['Daily_No_Ruc'] = master_df['Daily_No_Ruc'] * 100
    master_df['Daily'] = master_df['Daily'] * 100
    # format Zone according to email sent last evening
    # format Load according to email sent last evening

    # Find the first blank row
    first_blank_row = find_first_blank_row(ws_dst)
    if first_blank_row == 0:
        print("ERROR: Could not determine where to append data in master table")
        return 0

    # Get the next ID number
    next_id = get_next_id(ws_dst)

    # Append data rows
    rows_appended = 0
    write_row = first_blank_row

    for r_offset, row in enumerate(master_df.itertuples(index=False), start=0):
        dst_row = write_row + r_offset

        # Column A: ID (skip the ID from the DataFrame, use our own sequence)
        ws_dst.cell(row=dst_row, column=1, value=next_id + r_offset)

        # Apply master formats to the entire row first
        apply_master_formats(ws_dst, dst_row)

        # Columns B-Q: Write the master data
        # The row tuple has all columns including ID, so we need to map correctly
        row_dict = row._asdict()  # Convert named tuple to dict for easier access

        for col_idx, header in enumerate(MASTER_HEADERS):
            if header in row_dict:
                value = row_dict[header]
                ws_dst.cell(row=dst_row, column=col_idx + 2, value=value)  # +2 because B=2, C=3, etc.

        rows_appended += 1

    # Save the workbook
    wb_dst.save(dst_master_path)
    wb_dst.close()

    return rows_appended

# Backward-compatible alias for appending a master-formatted DataFrame
# Some callers expect append_master_formatted_dataframe_to_master
# Use existing implementation in a()

def append_master_formatted_dataframe_to_master(master_df: 'pd.DataFrame', dst_master: Path) -> int:
    return a(master_df, dst_master)




def write_updated_master_copy(master_df: 'pd.DataFrame',
                               master_dir: Path | str = Path('2-copy-reformat'),
                               master_filename: str = 'Master-Table.xlsx',
                               out_filename: str = 'master-file-updated.xlsx') -> Path:
    """Append master_df to the Master-Table but save as a new file without modifying the original.

    - Reads master from master_dir/master_filename
    - Computes next ID as (max Column A) + 1
    - Renumbers master_df['ID'] starting at that next ID
    - Appends rows to a workbook copy and saves to master_dir/out_filename
    - Returns the output path
    """
    from openpyxl import load_workbook
    import pandas as pd

    # Resolve paths
    master_dir = Path(master_dir)
    master_path = master_dir / master_filename
    out_path = master_dir / out_filename

    if not master_path.exists():
        raise FileNotFoundError(f"Master table not found: {master_path}")

    if master_df is None or master_df.empty:
        print("No data to append: input DataFrame is empty.")
        return out_path

    # Load the master workbook (do not modify original file)
    wb_dst = load_workbook(master_path)
    ws_dst = wb_dst.active

    # Determine write position and starting ID
    first_blank_row = find_first_blank_row(ws_dst)
    next_id = get_next_id(ws_dst)

    # Renumber the DataFrame's ID starting at next_id
    master_df = master_df.copy()
    master_df['ID'] = range(next_id, next_id + len(master_df))

    # Apply rows into the copy
    rows_appended = 0
    for r_offset, row in enumerate(master_df.itertuples(index=False), start=0):
        dst_row = first_blank_row + r_offset

        # Column A: write ID from the DataFrame
        apply_master_formats(ws_dst, dst_row)
        id_value = getattr(row, 'ID') if hasattr(row, 'ID') else (next_id + r_offset)
        ws_dst.cell(row=dst_row, column=1, value=id_value)

        # Columns B..Q in the MASTER_HEADERS order
        row_dict = row._asdict()
        for col_idx, header in enumerate(MASTER_HEADERS):
            if header in row_dict:
                ws_dst.cell(row=dst_row, column=col_idx + 2, value=row_dict[header])

        rows_appended += 1

    # Save to a new file path (do not overwrite original master)
    wb_dst.save(out_path)
    wb_dst.close()

    print(f"Wrote updated master copy: {out_path} ({rows_appended} rows appended)")
    return out_path




def hda_matrix_to_master_cols(df: 'pd.DataFrame') -> 'pd.DataFrame':
    """Transform HDA 'Matrix Table' sheet columns into master table schema.
    Attempts to be resilient to occasional header changes.

    Expected semantic fields:
      - Description of zone + load factor (e.g., MatrixDescription)
      - Price (base price column)
      - Term in months (e.g., TermCode)
      - Start date (e.g., StartDate)
      - Optional TDSP/Utility code
    Produces DataFrame with columns matching the master table structure.
    If a 'Product' column exists, keeps only rows where Product == 'Fixed Price'.
    """
    # Build normalization map for flexible header matching
    raw_cols = list(df.columns)
    def norm(s: str) -> str:
        return re.sub(r"[^a-z0-9]", "", str(s).strip().lower())
    norm_map = {norm(c): c for c in raw_cols}

    def get_col(candidates: list[str], allow_contains: bool = False):
        # Try exact normalized matches first
        for cand in candidates:
            key = norm(cand)
            if key in norm_map:
                return norm_map[key]
        if allow_contains:
            # Try contains on normalized keys
            for key_norm, orig in norm_map.items():
                for cand in candidates:
                    if norm(cand) in key_norm:
                        return orig
        return None

    # Candidate names for each logical field
    desc_candidates = ['MatrixDescription', 'Matrix Description', 'Description', 'Desc']
    price_candidates = ['Price', 'Price $', 'Price($)', 'Matrix Price', 'Rate', 'Base Price']
    green_candidates = ['GreenPrice', 'Green Price']
    term_candidates = ['TermCode', 'Term', 'Term Months', 'Term_Months', 'TermLength', 'Term (months)']
    start_candidates = ['StartDate', 'Start Date', 'StartMonth', 'Start Month', 'Delivery Start', 'First Delivery', 'DeliveryStart']
    created_candidates = ['CreatedDate', 'Created Date', 'Created', 'Price Date', 'PriceDate']
    tdsp_candidates = ['TdspCode', 'TDSP', 'TDSP Code', 'Utility']
    zone_candidates = ['Zone', 'Congestion Zone']
    lf_candidates = ['Load Factor', 'LoadFactor', 'LF']
    product_candidates = ['Product', 'Products']

    # Resolve columns
    c_desc = get_col(desc_candidates, allow_contains=True)
    c_price = get_col(price_candidates, allow_contains=True)
    c_green = get_col(green_candidates, allow_contains=True)
    c_term = get_col(term_candidates, allow_contains=True)
    c_start = get_col(start_candidates, allow_contains=True)
    c_created = get_col(created_candidates, allow_contains=True)
    c_tdsp = get_col(tdsp_candidates, allow_contains=True)
    c_zone = get_col(zone_candidates, allow_contains=True)
    c_lf = get_col(lf_candidates, allow_contains=True)
    c_prod = get_col(product_candidates, allow_contains=True)

    # Master table column structure
    master_cols = ['ID', 'Price_Date', 'Date', 'Zone', 'Load', 'REP1', 'Term', 'Min_MWh', 'Max_MWh',
                   'Daily_No_Ruc', 'RUC_Nodal', 'Daily', 'Com_Disc', 'HOA_Disc', 'Broker_Fee', 'Meter_Fee', 'Max_Meters']

    out = pd.DataFrame(columns=master_cols)

    # If we have a product column, enforce Product == 'Fixed Price'
    work_df = df
    if c_prod is not None:
        mask_fp = work_df[c_prod].astype(str).str.strip().str.lower().eq('fixed price')
        work_df = work_df[mask_fp]

    # Require core fields: price (or green price), term, start date; and desc or (zone+lf)
    if (c_price is None and c_green is None) or c_term is None or c_start is None:
        return out
    if c_desc is None and (c_zone is None and c_lf is None):
        return out

    # Get the number of rows we'll be working with
    num_rows = len(work_df)
    if num_rows == 0:
        return out

    # Helpers to derive zone and load factor from description
    def parse_zone(text: str) -> str:
        s = str(text or '').strip()
        for token in [' Low Load Factor', ' Medium Load Factor', ' High Load Factor']:
            if s.endswith(token):
                return s[: -len(token)]
        return s

    def parse_lf(text: str) -> str:
        print(f"parse_lf: {text}")
        s = str(text or '').strip()
        if 'Low Load Factor' in s:
            return 'LOW'
        elif 'Medium Load Factor' in s:
            return 'MED'
        elif 'High Load Factor' in s:
            return 'HIGH'
        return ''

    # Map to master table columns
    # Create DataFrame with proper index to avoid scalar assignment issues
    out = pd.DataFrame(index=range(num_rows), columns=master_cols)
    out['ID'] = None  # Will be set during append function

    # Price_Date should be today's date, Date should be start date
    from datetime import date as date_today
    today = date_today.today()
    out['Price_Date'] = today  # Column B - today's date for all rows

    start_dates = pd.to_datetime(work_df[c_start], errors='coerce')
    out['Date'] = start_dates.dt.date  # Column C - start date from input

    # Zone and Load Factor
    if c_desc is not None:
        out['Zone'] = work_df[c_desc].map(parse_zone)
        out['Load'] = work_df[c_desc].map(parse_lf)
    else:
        out['Zone'] = work_df[c_zone].astype(str) if c_zone is not None else ''
        out['Load'] = work_df[c_lf].astype(str) if c_lf is not None else ''

    # REP1 - hardcode to HUDSON to match 2-mapping.py
    out['REP1'] = 'HUDSON'

    # Term integer months with TARGET_TERMS filter
    def term_to_int(v):
        try:
            iv = int(float(v))
            return iv if iv in TARGET_TERMS else None
        except Exception:
            m = re.search(r"(\d+)", str(v))
            if m:
                iv = int(m.group(1))
                return iv if iv in TARGET_TERMS else None
            return None
    out['Term'] = work_df[c_term].map(term_to_int)

    # Usage tiers - match 2-mapping.py exactly
    out['Min_MWh'] = 0
    out['Max_MWh'] = 1000

    # Price columns - convert from cents to $/MWh (multiply by 10) to match 2-mapping.py
    price_series = pd.to_numeric(work_df[c_price], errors='coerce') if c_price is not None else pd.Series(dtype='float64')
    if price_series.isna().all() and c_green is not None:
        price_series = pd.to_numeric(work_df[c_green], errors='coerce')

    # Use price values as-is (should be around 70-80)
    # No multiplication needed if source values are already in correct range

    out['Daily_No_Ruc'] = price_series
    out['RUC_Nodal'] = 0  # Set to 0 to match 2-mapping.py
    out['Daily'] = price_series  # Same as Daily_No_Ruc

    # Default values for other columns - match 2-mapping.py exactly
    out['Com_Disc'] = 0
    out['HOA_Disc'] = 0
    out['Broker_Fee'] = 0
    out['Meter_Fee'] = 0
    out['Max_Meters'] = 5

    # Keep only rows with allowed terms
    out = out[out['Term'].notna()].copy()

    return out

def hda_matrix_to_base_cols(df: 'pd.DataFrame') -> 'pd.DataFrame':
    """Legacy function - now calls the new master table function for compatibility."""
    return hda_matrix_to_master_cols(df)


# Robust transformer v2: handle header variants and map Matrix Table -> BASE_COLS

def hda_matrix_to_base_cols_v2(df: 'pd.DataFrame') -> 'pd.DataFrame':
    raw_cols = list(df.columns)
    def norm(s: str) -> str:
        return re.sub(r"[^a-z0-9]", "", str(s).strip().lower())
    # Build mapping from normalized -> original
    norm_map = {norm(c): c for c in raw_cols}

    def get_col(candidates: list[str], allow_contains: bool = False):
        for cand in candidates:
            key = norm(cand)
            if key in norm_map:
                return norm_map[key]
        if allow_contains:
            for key_norm, orig in norm_map.items():
                for cand in candidates:
                    if norm(cand) in key_norm:
                        return orig
        return None

    # Candidate sets
    desc_candidates = ['MatrixDescription','Matrix Description','Description','Desc']
    price_candidates = ['Price','Price $','Price($)','Matrix Price','Rate','Base Price']
    green_candidates = ['GreenPrice','Green Price']
    term_candidates  = ['TermCode','Term','Term Months','Term_Months','TermLength','Term (months)']
    start_candidates = ['StartDate','Start Date','StartMonth','Start Month','Delivery Start','First Delivery','DeliveryStart']
    tdsp_candidates  = ['TdspCode','TDSP','TDSP Code','Utility']
    zone_candidates  = ['Zone','Congestion Zone']
    lf_candidates    = ['Load Factor','LoadFactor','LF']
    product_candidates = ['Product','Products']

    c_desc  = get_col(desc_candidates, allow_contains=True)
    c_price = get_col(price_candidates, allow_contains=True)
    c_green = get_col(green_candidates, allow_contains=True)
    c_term  = get_col(term_candidates, allow_contains=True)
    c_start = get_col(start_candidates, allow_contains=True)
    c_tdsp  = get_col(tdsp_candidates, allow_contains=True)
    c_zone  = get_col(zone_candidates, allow_contains=True)
    c_lf    = get_col(lf_candidates, allow_contains=True)
    c_prod  = get_col(product_candidates, allow_contains=True)

    out = pd.DataFrame(columns=BASE_COLS)

    # Optional pre-filter Product == Fixed Price
    work_df = df
    if c_prod is not None:
        mask_fp = work_df[c_prod].astype(str).str.strip().str.lower().eq('fixed price')
        work_df = work_df[mask_fp]

    # Require essentials
    if c_term is None or c_start is None:
        return out
    if c_price is None and c_green is None:
        return out
    if c_desc is None and (c_zone is None and c_lf is None):
        return out

    def parse_zone(text: str) -> str:
        s = str(text or '').strip()
        for token in [' Low Load Factor',' Medium Load Factor',' High Load Factor']:
            if s.endswith(token):
                return s[:-len(token)]
        return s

    def parse_lf(text: str) -> str:
        s = str(text or '').strip()
        if 'Low Load Factor' in s:
            return 'LOW'
        elif 'Medium Load Factor' in s:
            return 'MED'
        elif 'High Load Factor' in s:
            return 'HIGH'
        return ''

    def term_to_int(v):
        try:
            iv = int(float(v))
            return iv if iv in TARGET_TERMS else None
        except Exception:
            m = re.search(r"(\d+)", str(v))
            if m:
                iv = int(m.group(1))
                return iv if iv in TARGET_TERMS else None
            return None

    out['Start Month'] = pd.to_datetime(work_df[c_start], errors='coerce').dt.date
    out['State'] = 'TX'
    out['Utility'] = work_df[c_tdsp].astype(str) if c_tdsp is not None else ''

    if c_desc is not None:
        out['Congestion Zone'] = work_df[c_desc].map(parse_zone)
        out['Load Factor'] = work_df[c_desc].map(parse_lf)
    else:
        out['Congestion Zone'] = work_df[c_zone].astype(str) if c_zone is not None else ''
        out['Load Factor'] = work_df[c_lf].astype(str) if c_lf is not None else ''

    out['Term'] = work_df[c_term].map(term_to_int)
    out['Product'] = 'Fixed Price'

    price_series = pd.to_numeric(work_df[c_price], errors='coerce') if c_price is not None else pd.Series(dtype='float64')
    if price_series.isna().all() and c_green is not None:
        price_series = pd.to_numeric(work_df[c_green], errors='coerce')
    out['0-200,000'] = price_series

    out = out[out['Term'].notna()].copy()
    out = out[BASE_COLS]
    return out


def create_template_copy_with_filtered_values(src_xlsm: Path,
                                               template_path: Path,
                                               out_copy_path: Path,
                                               sheet_name_prefer: str = 'matrix table') -> Path | None:
    """Implements the detailed Template step:
    - Make a new Template copy (from template_path to out_copy_path)
    - Read filtered values from src_xlsm (Matrix Table -> BASE_COLS via transformer)
    - Paste those values into the copy at A..H starting at row 2
    - Leave formula columns as-is
    Returns the path to the created copy, or None on failure.
    """
    if not src_xlsm.exists():
        print(f"ERROR: Source not found: {src_xlsm}")
        return None
    if not template_path.exists():
        print(f"ERROR: Template not found: {template_path}")
        return None

    # Read Matrix Table and transform to BASE_COLS
    try:
        all_sheets = pd.read_excel(src_xlsm, sheet_name=None)
    except Exception as e:
        print('READ_ERROR')
        print(str(e))
        return None

    # Resolve sheet name
    target = None
    wanted = sheet_name_prefer.strip().lower()
    for name in all_sheets.keys():
        if str(name).strip().lower() == wanted:
            target = name
            break
    if target is None:
        words = [w for w in wanted.split() if w]
        for name in all_sheets.keys():
            nrm = str(name).strip().lower()
            if all(w in nrm for w in words):
                target = name
                break
    if target is None:
        print(f"WARNING: Preferred sheet '{sheet_name_prefer}' not found in {src_xlsm.name}.")
        return None

    transformed = hda_matrix_to_master_cols(all_sheets[target])
    if transformed.empty:
        print("No rows after transformation/filtering; template copy not created.")
        return None

    # Make the new template copy
    try:
        shutil.copyfile(template_path, out_copy_path)
    except Exception as e:
        print(f"ERROR copying template: {e}")
        return None

    # Paste A..H values starting at row 2 in the copy's first sheet
    wb = load_workbook(out_copy_path)
    ws = wb.worksheets[0]

    # Write headers into row 1 (optional, keeps template headers if present)
    for c, col_name in enumerate(BASE_COLS, start=1):
        ws.cell(row=1, column=c, value=col_name)

    # Write values
    for r_offset, row in enumerate(transformed.itertuples(index=False), start=0):
        r = 2 + r_offset
        for c_offset, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c_offset, value=value)
            if isinstance(value, (datetime, date)):
                cell.number_format = 'm/dd/yyyy'

    try:
        wb.save(out_copy_path)
    except PermissionError:
        print("ERROR: Could not save template copy. Please close it if it's open and re-run.")
        return None

    print(f"Template copy created: {out_copy_path}")
    return out_copy_path

# Write HDA filtered output analogous to 'ERCOT-filtered.xlsx'

def write_hda_filtered(src_xlsm: Path,
                       out_path: Path,
                       sheet_name_prefer: str = 'matrix table') -> int:
    """Create an 'HDA-filtered.xlsx' style file from the HDA .xlsm Matrix Table.
    Returns number of rows written.
    """
    if not src_xlsm.exists():
        print(f"ERROR: Source not found: {src_xlsm}")
        return 0

    try:
        all_sheets = pd.read_excel(src_xlsm, sheet_name=None)
    except Exception as e:
        print('READ_ERROR')
        print(str(e))
        return 0

    # Resolve sheet name
    target = None
    wanted = sheet_name_prefer.strip().lower()
    for name in all_sheets.keys():
        if str(name).strip().lower() == wanted:
            target = name
            break
    if target is None:
        words = [w for w in wanted.split() if w]
        for name in all_sheets.keys():
            nrm = str(name).strip().lower()
            if all(w in nrm for w in words):
                target = name
                break
    if target is None:
        print(f"WARNING: Preferred sheet '{sheet_name_prefer}' not found in {src_xlsm.name}.")
        return 0

    transformed = hda_matrix_to_master_cols(all_sheets[target])
    if transformed.empty:
        print("No rows after transformation/filtering.")
        return 0

    # Write to Excel analogous to ERCOT-filtered.xlsx
    try:
        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            transformed.to_excel(writer, sheet_name='Sheet1', index=False)
    except Exception as e:
        print('WRITE_ERROR')
        print(str(e))
        return 0

    print(f"Wrote HDA-filtered output: {out_path} ({len(transformed)} rows)")
    return len(transformed)

def process_xlsm_file(src_xlsm: Path, dst_master: Path, sheet_name_prefer: str | None = None) -> int:
    """Process a single .xlsm file: read the requested sheet (if provided) or all sheets,
    apply filter_sheet, and append to master.
    """
    if not src_xlsm.exists():
        print(f"ERROR: Source file not found: {src_xlsm}")
        return 0

    if not dst_master.exists():
        print(f"ERROR: Master table not found: {dst_master}")
        return 0

    try:
        # Read the .xlsm file and process it using HDA matrix transformation
        all_sheets = pd.read_excel(src_xlsm, sheet_name=None)

        # Find the matrix table sheet
        target_sheet = None
        if sheet_name_prefer:
            # Look for the preferred sheet name
            wanted = sheet_name_prefer.strip().lower()
            for name in all_sheets.keys():
                if str(name).strip().lower() == wanted:
                    target_sheet = name
                    break
            if target_sheet is None:
                words = [w for w in wanted.split() if w]
                for name in all_sheets.keys():
                    nrm = str(name).strip().lower()
                    if all(w in nrm for w in words):
                        target_sheet = name
                        break
        else:
            # Default to looking for 'matrix table' sheet
            for name in all_sheets.keys():
                if 'matrix' in str(name).lower() and 'table' in str(name).lower():
                    target_sheet = name
                    break

        if target_sheet is None:
            print(f"WARNING: No suitable sheet found in {src_xlsm.name}. Available sheets: {list(all_sheets.keys())}")
            return 0

        print(f"line 1092 Processing sheet '{target_sheet}' from {src_xlsm.name}")

        # Transform the data using HDA matrix transformation
        print(f"line 1099 Processing sheet '{target_sheet}' from {src_xlsm.name}")
        transformed_df = hda_matrix_to_master_cols(all_sheets[target_sheet])

        if transformed_df.empty:
            print(f"No data to append after transformation from {src_xlsm.name}")
            return 0

        # Append to master table using the master-formatted DataFrame
        rows_appended = append_master_formatted_dataframe_to_master(transformed_df, dst_master)
        print(f"Appended {rows_appended} rows from {src_xlsm.name} to {dst_master.name}")
        return rows_appended

    except Exception as e:
        print(f"ERROR processing {src_xlsm.name}: {str(e)}")
        return 0

# === Header-mapped append (from nice-scripts pattern) ===
# This appends rows from a template workbook by header names and applies
# master number formats so pasted cells match the master table's formatting.

from typing import Dict, List

# Expected master headers in order (B..Q)
MASTER_HEADERS: List[str] = [
    'Price_Date','Date','Zone','Load','REP1','Term','Min_MWh','Max_MWh',
    'Daily_No_Ruc','RUC_Nodal','Daily','Com_Disc','HOA_Disc','Broker_Fee','Meter_Fee','Max_Meters'
]

# Source header normalization map (case-insensitive)
SOURCE_SYNONYMS: Dict[str, str] = {
    'price_date': 'Price_Date',
    'date': 'Date',
    'zone': 'Zone',
    'rep1': 'REP1',
    'load': 'Load',
    'term': 'Term',
    'min_mwh': 'Min_MWh',
    'max_mwh': 'Max_MWh',
    'daily_no_ruc': 'Daily_No_Ruc',
    'ruc_nodal': 'RUC_Nodal',  # normalize template's Ruc_Nodal
    'ruc_nodal_': 'RUC_Nodal',
    'ruc_nodal__': 'RUC_Nodal',
    'daily': 'Daily',
    'com_disc': 'Com_Disc',
    'hoa_disc': 'HOA_Disc',
    'broker_fee': 'Broker_Fee',
    'meter_fee': 'Meter_Fee',
    'max_meters': 'Max_Meters',
}

# Master number formats for A..Q (IDs included)
MASTER_FORMATS: Dict[str, str] = {
    'A': 'General',
    'B': 'mm-dd-yy',
    'C': 'mm-dd-yy',
    'D': 'General',
    'E': 'General',
    'F': 'General',
    'G': '* #,##0;* (#,##0);* -00',
    'H': '* #,##0;* (#,##0);* -00',
    'I': 'General',
    'J': '$#,##0.00;($#,##0.00)',
    'K': '$#,##0.00',
    'L': '$#,##0.00',
    'M': '$* #,##0.00;$* (#,##0.00);$* -00',
    'N': '$* #,##0.00;$* (#,##0.00);$* -00',
    'O': '$#,##0.00',
    'P': '$* #,##0.00;$* (#,##0.00);$* -00',
    'Q': '#,##0',
}


def detect_header_row(ws, max_scan_rows: int = 30) -> int:
    """Find the header row by scanning for expected header names."""
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        norm = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                norm.append(v.strip().lower().replace(' ', '_'))
            else:
                norm.append(None)
        if 'price_date' in norm and ('zone' in norm or 'load' in norm) and ('daily' in norm or 'daily_no_ruc' in norm):
            return r
    return 1


def build_source_mapping(ws, header_row: int) -> Dict[str, int]:
    """Map master header names to source column indices by header text."""
    mapping: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        hv = ws.cell(row=header_row, column=c).value
        if not isinstance(hv, str):
            continue
        key = hv.strip().lower().replace(' ', '_')
        if key in SOURCE_SYNONYMS:
            norm_name = SOURCE_SYNONYMS[key]
            if norm_name in MASTER_HEADERS and norm_name not in mapping:
                mapping[norm_name] = c
    return mapping


def apply_master_formats(ws_dst, row_idx: int) -> None:
    """Apply master number formats for columns A..Q to a single row."""
    from openpyxl.styles import Alignment

    for c in range(1, 17 + 1):
        col_letter = get_column_letter(c)
        fmt = MASTER_FORMATS.get(col_letter)
        cell = ws_dst.cell(row=row_idx, column=c)
        if fmt:
            cell.number_format = fmt

        # Apply right alignment to REP1 column (column F = 6)
        if c == 6:  # Column F (REP1/HUDSON)
            cell.alignment = Alignment(horizontal='right')


def append_from_template(template_path: Path, template_sheet: str, master_path: Path) -> None:
    """Append rows from a template workbook to the master by header mapping and apply number formats."""
    # Create backup before modifying master table
    backup_path = create_master_table_backup(master_path)
    if backup_path is None:
        print("Warning: Could not create backup, proceeding anyway...")
    else:
        print(f"Master table backed up to: {backup_path}")

    # Load source workbook
    wb_src = load_workbook(template_path, data_only=True, read_only=False)
    if template_sheet not in wb_src.sheetnames:
        raise ValueError(f"Sheet '{template_sheet}' not found in template: {template_path}")
    ws_src = wb_src[template_sheet]

    # Discover header row and build mapping
    header_row = detect_header_row(ws_src)
    src_map = build_source_mapping(ws_src, header_row)
    missing = [h for h in MASTER_HEADERS if h not in src_map]
    if missing:
        raise ValueError('Missing expected columns in template: ' + ', '.join(missing))

    # Open master workbook
    wb_dst = load_workbook(master_path)
    # First visible sheet
    ws_dst = None
    for name in wb_dst.sheetnames:
        sh = wb_dst[name]
        if getattr(sh, 'sheet_state', 'visible') == 'visible':
            ws_dst = sh
            break
    if ws_dst is None:
        ws_dst = wb_dst.active

    # Determine current max ID in column A
    max_id = 0
    for r in range(2, ws_dst.max_row + 1):
        v = ws_dst.cell(row=r, column=1).value
        if isinstance(v, (int, float)):
            try:
                vi = int(v)
            except Exception:
                continue
            if vi > max_id:
                max_id = vi
    next_id = max_id + 1

    # Find the first blank ID row (column A)
    first_blank_row = None
    for r in range(2, ws_dst.max_row + 1):
        v = ws_dst.cell(row=r, column=1).value
        if v in (None, ''):
            first_blank_row = r
            break
    if first_blank_row is None:
        first_blank_row = ws_dst.max_row + 1

    # Iterate source rows and append
    start_data_row = header_row + 1
    rows_appended = 0
    write_row = first_blank_row

    for r in range(start_data_row, ws_src.max_row + 1):
        values: List[object] = []
        all_empty = True
        for h in MASTER_HEADERS:
            c_idx = src_map[h]
            cell = ws_src.cell(row=r, column=c_idx)
            v = cell.value
            if isinstance(v, str):
                v = v.strip()
            if v not in (None, ''):
                all_empty = False
            values.append(v)
        if all_empty:
            continue

        dst_row = write_row
        # ID
        ws_dst.cell(row=dst_row, column=1, value=next_id)
        # Apply formats to A..Q first
        apply_master_formats(ws_dst, dst_row)
        # Write values for B..Q in the master-defined order
        for i, val in enumerate(values):
            ws_dst.cell(row=dst_row, column=i + 2, value=val)

        write_row += 1
        rows_appended += 1
        next_id += 1

    if rows_appended == 0:
        print('No non-empty rows found to append.')
    else:
        wb_dst.save(master_path)
        print(f'Appended {rows_appended} rows by header mapping. IDs {max_id + 1}..{next_id - 1}.')


def print_usage():
    """Print usage information for the excel processor"""
    print("Excel Processor with SharePoint Integration")
    print("Usage:")
    print("  python excel_processor.py download-sharepoint <filename> [<master-table-path>] [<sheet-name>]")
    print("    Download and process a file from SharePoint")
    print("  python excel_processor.py download-only <filename> [<local-path>]")
    print("    Download a file from SharePoint without processing")
    print("  python excel_processor.py process-hda <path-to-file-or-dir> [<master-table-path>]")
    print("    Process HDA .xlsm files")
    print("  python excel_processor.py append-l-aa <source-path> [<master-table-path>]")
    print("    Append L..AA columns from source to master table")
    print("  python excel_processor.py append-from-template <template-path> <sheet-name> [<master-table-path>]")
    print("    Append from template by header mapping")
    print("  python excel_processor.py")
    print("    Default: process unfiltered source into master table")
    print()
    print("SharePoint Configuration:")
    print("  Ensure .env file contains: TENANT_ID, CLIENT_ID, CLIENT_SECRET, SITE_HOSTNAME, SITE_PATH, SHAREPOINT_UPLOAD_FOLDER")


def main():
    root = Path('.')

    # Help mode
    if len(sys.argv) >= 2 and sys.argv[1] in ['--help', '-h', 'help']:
        print_usage()
        return

    # SharePoint download and process mode
    if len(sys.argv) >= 3 and sys.argv[1] == 'download-sharepoint':
        file_name = sys.argv[2]
        dst_arg = Path(sys.argv[3]) if len(sys.argv) >= 4 else Path(DST_MASTER_TABLE_NAME)
        sheet_name = sys.argv[4] if len(sys.argv) >= 5 else None

        # The function will handle downloading the master table if it doesn't exist
        rows_appended = download_and_process_sharepoint_file(file_name, dst_arg, sheet_name)
        if rows_appended and rows_appended > 0:
            print(f"SUCCESS: Downloaded and processed {file_name} from SharePoint. {rows_appended} rows appended to {dst_arg.name}.")
        else:
            print(f"No rows were appended from {file_name}.")
        return

    # SharePoint download only (no processing)
    if len(sys.argv) >= 3 and sys.argv[1] == 'download-only':
        file_name = sys.argv[2]
        download_path = sys.argv[3] if len(sys.argv) >= 4 else None

        downloaded_file = download_sharepoint_file(file_name, download_path)
        if downloaded_file:
            print(f"SUCCESS: Downloaded {file_name} to {downloaded_file}")
        else:
            print(f"FAILED: Could not download {file_name} from SharePoint")
        return

    # If called with explicit args (append mode), run append_l_aa and exit
    if len(sys.argv) >= 3 and sys.argv[1] == 'append-l-aa':
        src_arg = Path(sys.argv[2])
        dst_arg = Path(sys.argv[3]) if len(sys.argv) >= 4 else Path(DST_MASTER_TABLE_NAME)
        append_l_aa(src_arg, dst_arg)
        return

    # Standalone: append from a template workbook into the master by header mapping
    if len(sys.argv) >= 4 and sys.argv[1] == 'append-from-template':
        template_arg = Path(sys.argv[2])
        sheet_arg = sys.argv[3]
        master_arg = Path(sys.argv[4]) if len(sys.argv) >= 5 else Path(DST_MASTER_TABLE_NAME)
        if not template_arg.exists():
            print(f"ERROR: Template file not found: {template_arg}")
            return
        if not master_arg.exists():
            print(f"ERROR: Master file not found: {master_arg}")
            return
        try:
            append_from_template(template_arg, sheet_arg, master_arg)
        except Exception as e:
            print('APPEND_FROM_TEMPLATE_ERROR')
            print(str(e))
        return


    # If called with 'process-hda', process a .xlsm file or all .xlsm files in a directory
    if len(sys.argv) >= 2 and sys.argv[1] == 'process-hda':
        # Allow omitting the path to default to a folder named 'HDA' if present
        target_path = Path(sys.argv[2]) if len(sys.argv) >= 3 else Path('HDA')
        dst_arg = Path(sys.argv[3]) if len(sys.argv) >= 4 else Path(DST_MASTER_TABLE_NAME)

        if not target_path.exists():
            print(f"ERROR: HDA path not found: {target_path}")
            print("Usage: python excel_processor.py process-hda <path-to-file-or-dir> [<dst-master-path>]")
            return

        appended_total = 0
        processed_files = []

        if target_path.is_file() and target_path.suffix.lower() == '.xlsm':
            appended_total += process_xlsm_file(target_path, dst_arg, sheet_name_prefer='matrix table')
            processed_files.append(target_path)
        elif target_path.is_dir():
            xlsm_files = list(target_path.glob('*.xlsm'))
            if not xlsm_files:
                # Fall back to recursive search
                xlsm_files = list(target_path.rglob('*.xlsm'))
            if not xlsm_files:
                print(f"No .xlsm files found under: {target_path}")
                return
            for f in sorted(xlsm_files):
                appended_total += process_xlsm_file(f, dst_arg, sheet_name_prefer='matrix table')
                processed_files.append(f)
        else:
            print(f"ERROR: Unsupported HDA target: {target_path}")
            print("Provide a .xlsm file or a directory containing .xlsm files.")
            return

        print(f"HDA processing complete. Files processed: {len(processed_files)}; Rows appended: {appended_total}.")
        for f in processed_files:
            print(f" - {f}")
        return

    # Default behavior (legacy): read and filter unfiltered source into master table, then add formulas
    src = root / SRC_UNFILTERED_NAME
    dst = root / DST_MASTER_TABLE_NAME

    if not src.exists():
        print(f"ERROR: Unfiltered source file not found: {src}")
        sys.exit(3)
    if not dst.exists():
        print(f"ERROR: Master table file not found: {dst}")
        sys.exit(3)

    # Append rows from a template workbook into the master by header mapping (standalone mode)
    if len(sys.argv) >= 4 and sys.argv[1] == 'append-from-template':
        template_arg = Path(sys.argv[2])
        sheet_arg = sys.argv[3]
        master_arg = Path(sys.argv[4]) if len(sys.argv) >= 5 else Path(DST_MASTER_TABLE_NAME)
        if not template_arg.exists():
            print(f"ERROR: Template file not found: {template_arg}")
            return
        if not master_arg.exists():
            print(f"ERROR: Master file not found: {master_arg}")
            return
        try:
            append_from_template(template_arg, sheet_arg, master_arg)
        except Exception as e:
            print('APPEND_FROM_TEMPLATE_ERROR')
            print(str(e))
        return

    # 1. Read and filter the source data (legacy default path)
    try:
        sheets = pd.read_excel(src, sheet_name=None)
    except Exception as e:
        print('READ_ERROR')
        print(str(e))
        sys.exit(2)

    filtered_data = []
    for _, df in sheets.items():
        filtered_data.append(filter_sheet(df))

    combined_df = pd.concat(filtered_data, ignore_index=True)

    if combined_df.empty:
        print("No data to append after filtering.")
        return

    # 2. Load the master table workbook
    wb_dst = load_workbook(dst)
    ws_dst = wb_dst.active

    last_row = last_data_row(ws_dst)
    start_row = last_row + 1 if last_row >= 1 else 1

    # 3. Capture the formatting from the last existing row in the master table
    dst_formats = {}
    if last_row > 0:
        for col_idx in range(1, len(BASE_COLS) + 1):
            dst_formats[col_idx] = ws_dst.cell(row=last_row, column=col_idx).number_format
    else:
        # If the master table is empty, default to 'General' format
        for col_idx in range(1, len(BASE_COLS) + 1):
            dst_formats[col_idx] = 'General'

    # 4. Append filtered data and apply formatting
    for r_offset, row_data in enumerate(combined_df.itertuples(index=False), start=0):
        r = start_row + r_offset

        # This part handles the automatic row numbering in column A
        if r == 1:
            ws_dst.cell(row=r, column=1, value=1)
        else:
            prev_val = ws_dst.cell(row=r-1, column=1).value
            try:
                val_int = int(prev_val) if prev_val is not None and str(prev_val).strip() != '' else 0
            except (ValueError, TypeError):
                val_int = 0
            ws_dst.cell(row=r, column=1, value=val_int + 1)

        # Paste the values from the filtered data and apply formatting
        for c_offset, value in enumerate(row_data, start=1):
            # This handles the column mapping from BASE_COLS to the master table
            cell = ws_dst.cell(row=r, column=c_offset + 1, value=value)
            cell.number_format = dst_formats.get(c_offset + 1, 'General')

            # Special handling for dates
            if isinstance(value, datetime) or isinstance(value, date):
                cell.number_format = 'm/dd/yyyy'

    end_row = start_row + len(combined_df) - 1

    # 5. Add formulas to the newly appended rows
    # The formulas are added to the columns to the right of the pasted data
    add_formulas(dst, start_row, end_row)

    try:
        wb_dst.save(dst)
        print(f"SUCCESS: Appended {len(combined_df)} rows and formulas to {DST_MASTER_TABLE_NAME}.")
    except PermissionError:
        print("ERROR: Could not save destination file. Please close it if it's open and re-run.")
        sys.exit(4)

if __name__ == '__main__':
    main()