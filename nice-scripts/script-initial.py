import os
import sys
from datetime import datetime, date
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

MASTER_PATH = 'DAILY PRICING - new.xlsx'
TEMPLATE_PATH = "1-ERCOT/DAILY PRICING - APGE - TEMPLATE - 2024'.xlsx"
TEMPLATE_SHEET = 'IMPORT'

# Expected master headers in order (B..Q)
master_headers = [
    'Price_Date','Date','Zone','REP1','Load','Term','Min_MWh','Max_MWh',
    'Daily_No_Ruc','RUC_Nodal','Daily','Com_Disc','HOA_Disc','Broker_Fee','Meter_Fee','Max_Meters'
]

# Source header normalization map (case-insensitive)
source_synonyms = {
    'price_date': 'Price_Date',
    'date': 'Date',
    'zone': 'Zone',
    'rep1': 'REP1',
    'load': 'Load',
    'term': 'Term',
    'min_mwh': 'Min_MWh',
    'max_mwh': 'Max_MWh',
    'daily_no_ruc': 'Daily_No_Ruc',
    'ruc_nodal': 'RUC_Nodal',  # normalize template's Ruc_Nodal
    'ruc_nodal_': 'RUC_Nodal',
    'ruc_nodal__': 'RUC_Nodal',
    'daily': 'Daily',
    'com_disc': 'Com_Disc',
    'hoa_disc': 'HOA_Disc',
    'broker_fee': 'Broker_Fee',
    'meter_fee': 'Meter_Fee',
    'max_meters': 'Max_Meters',
}

# Master number formats for A..Q (IDs included)
master_formats = {
    'A': 'General',
    'B': 'mm-dd-yy',
    'C': 'mm-dd-yy',
    'D': 'General',
    'E': 'General',
    'F': 'General',
    'G': '* #,##0;* (#,##0);* -00',
    'H': '* #,##0;* (#,##0);* -00',
    'I': 'General',
    'J': '$#,##0.00;($#,##0.00)',
    'K': '$#,##0.00',
    'L': '$#,##0.00',
    'M': '$* #,##0.00;$* (#,##0.00);$* -00',
    'N': '$* #,##0.00;$* (#,##0.00);$* -00',
    'O': '$#,##0.00',
    'P': '$* #,##0.00;$* (#,##0.00);$* -00',
    'Q': '#,##0',
}

# 1) Load template
try:
    wb_src = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True, read_only=False)
except Exception as e:
    print(f"ERROR: Unable to open template: {e}")
    sys.exit(2)

if TEMPLATE_SHEET not in wb_src.sheetnames:
    print(f"ERROR: Sheet '{TEMPLATE_SHEET}' not found in template.")
    sys.exit(2)
ws_src = wb_src[TEMPLATE_SHEET]

# 2) Discover header row: scan first 30 rows for one containing key headers
header_row = None
for r in range(1, min(ws_src.max_row, 30)+1):
    row_vals = [ws_src.cell(row=r, column=c).value for c in range(1, ws_src.max_column+1)]
    norm = []
    for v in row_vals:
        if isinstance(v, str):
            k = v.strip().lower().replace(' ', '_')
            norm.append(k)
        else:
            norm.append(None)
    if 'price_date' in norm and ('zone' in norm or 'load' in norm) and ('daily' in norm or 'daily_no_ruc' in norm):
        header_row = r
        break
if header_row is None:
    header_row = 1

# Map headers to columns
src_col_by_master = {}
header_values = {}
for c in range(1, ws_src.max_column+1):
    hv = ws_src.cell(row=header_row, column=c).value
    if not isinstance(hv, str):
        continue
    key = hv.strip().lower().replace(' ', '_')
    if key in source_synonyms:
        norm_name = source_synonyms[key]
        header_values[norm_name] = hv.strip()
        if norm_name in master_headers and norm_name not in src_col_by_master:
            src_col_by_master[norm_name] = c

missing = [h for h in master_headers if h not in src_col_by_master]
if missing:
    print('ERROR: Missing expected columns in template:', ', '.join(missing))
    sys.exit(2)

# 3) Open master workbook
try:
    wb_dst = openpyxl.load_workbook(MASTER_PATH)
except PermissionError:
    print('ERROR: Could not open master workbook for writing. Please close it in Excel and retry.')
    sys.exit(3)
except Exception as e:
    print(f"ERROR: Unable to open master workbook: {e}")
    sys.exit(3)

# first visible sheet
ws_dst = None
for name in wb_dst.sheetnames:
    sh = wb_dst[name]
    if getattr(sh, 'sheet_state', 'visible') == 'visible':
        ws_dst = sh
        break
if ws_dst is None:
    ws_dst = wb_dst.active

# 4) Determine current max ID in column A
max_id = 0
for r in range(2, ws_dst.max_row+1):
    v = ws_dst.cell(row=r, column=1).value
    if isinstance(v, (int, float)):
        try:
            vi = int(v)
        except Exception:
            continue
        if vi > max_id:
            max_id = vi
next_id = max_id + 1

# 5) Iterate source rows and collect values
start_data_row = header_row + 1
rows_appended = 0
preview = []

for r in range(start_data_row, ws_src.max_row+1):
    # Build the row values according to master_headers order
    values = []
    all_empty = True
    for h in master_headers:
        c_idx = src_col_by_master[h]
        cell = ws_src.cell(row=r, column=c_idx)
        v = cell.value
        if isinstance(v, str):
            v = v.strip()
        if v not in (None, ''):
            all_empty = False
        values.append(v)
    if all_empty:
        # skip empty rows
        continue

    # Append to destination: A=ID, B..Q = values in order
    dst_row = ws_dst.max_row + 1
    ws_dst.cell(row=dst_row, column=1, value=next_id)
    # Set formats and values for A..Q
    for i, col_letter in enumerate([get_column_letter(c) for c in range(1, 17+1)]):
        fmt = master_formats.get(col_letter)
        cell = ws_dst.cell(row=dst_row, column=i+1)
        if fmt:
            cell.number_format = fmt
    # write B..Q
    for i, val in enumerate(values):
        ws_dst.cell(row=dst_row, column=i+2, value=val)

    # simple preview capture for first 3 rows
    if rows_appended < 3:
        preview.append([ws_dst.cell(row=dst_row, column=c).value for c in range(1, 8)])

    rows_appended += 1
    next_id += 1

# 6) Save
if rows_appended == 0:
    print('No non-empty rows found to append.')
    sys.exit(0)

try:
    wb_dst.save(MASTER_PATH)
except PermissionError:
    print('ERROR: Could not save master workbook. Please close it in Excel and retry.')
    sys.exit(4)

print(f'Appended {rows_appended} rows. New ID range: {max_id+1}..{next_id-1}')
print('Preview of first 3 appended rows (A..G):')
for row in preview:
    print(row)
