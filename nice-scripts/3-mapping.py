from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd

master_path = Path('1-original-excel-data/DAILY PRICING - new.xlsx')
hudson_path = Path('HDA/HudsonMatrixPrices08272025020701PM.xlsm')

# Recompute how many rows we appended from Hudson "Matrix Table"
try:
    all_sheets = pd.read_excel(hudson_path, sheet_name=None)
    mt_name = next(n for n in all_sheets.keys() if str(n).strip().lower() == 'matrix table')
    appended_count = len(all_sheets[mt_name])
except Exception as e:
    print('ERROR: Could not read Hudson Matrix Table to compute appended count:', e)
    appended_count = None

wb = load_workbook(master_path)
# First visible sheet
ws = None
for name in wb.sheetnames:
    sh = wb[name]
    if getattr(sh, 'sheet_state', 'visible') == 'visible':
        ws = sh
        break
if ws is None:
    ws = wb.active

max_row = ws.max_row
# Determine the appended range
if appended_count is None:
    start_row = 2
else:
    previous_data = (max_row - 1) - appended_count
    start_row = previous_data + 2
end_row = max_row

MASTER_FORMATS = {
    'A': 'General',
    'B': 'mm-dd-yy',
    'C': 'mm-dd-yy',
    'D': 'General',
    'E': 'General',
    'F': 'General',
    'G': '* #,##0;* (#,##0);* -00',
    'H': '* #,##0;* (#,##0);* -00',
    'I': 'General',
    'J': '$#,##0.00;($#,##0.00)',
    'K': '$#,##0.00',
    'L': '$#,##0.00',
    'M': '$* #,##0.00;$* (#,##0.00);$* -00',
    'N': '$* #,##0.00;$* (#,##0.00);$* -00',
    'O': '$#,##0.00',
    'P': '$* #,##0.00;$* (#,##0.00);$* -00',
    'Q': '#,##0',
}

for r in range(start_row, end_row + 1):
    for c in range(1, 17 + 1):
        fmt = MASTER_FORMATS.get(get_column_letter(c))
        if fmt:
            ws.cell(row=r, column=c).number_format = fmt

wb.save(master_path)
print(f'Formatted rows {start_row}..{end_row} with master formats for A..Q.')