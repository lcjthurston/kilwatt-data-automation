#!/usr/bin/env python3
"""
Test script to verify the Load Factor and HUDSON alignment fixes
"""

import pandas as pd
from pathlib import Path
import excel_processor as ep

def test_load_factor_parsing():
    """Test that Load Factor parsing returns simplified values"""
    print("Testing Load Factor parsing...")
    
    # Test the parse_lf function from hda_matrix_to_master_cols
    test_descriptions = [
        "North Zone Low Load Factor",
        "South Zone Medium Load Factor", 
        "West Zone High Load Factor",
        "Some other text"
    ]
    
    # Create a simple test DataFrame
    df = pd.DataFrame({
        'MatrixDescription': test_descriptions,
        'Price': [50, 60, 70, 80],
        'TermCode': [12, 24, 36, 12],
        'StartDate': ['2024-01-01', '2024-02-01', '2024-03-01', '2024-04-01']
    })
    
    # Transform using the function
    result = ep.hda_matrix_to_master_cols(df)
    
    print("Load Factor values:")
    for i, lf in enumerate(result['Load']):
        print(f"  {test_descriptions[i]} -> {lf}")
    
    # Check if we get the expected simplified values
    expected = ['LOW', 'MED', 'HIGH', '']
    actual = list(result['Load'])
    
    if actual == expected:
        print("✓ Load Factor parsing works correctly!")
        return True
    else:
        print(f"✗ Load Factor parsing failed. Expected: {expected}, Got: {actual}")
        return False

def test_hudson_alignment():
    """Test that HUDSON alignment is applied"""
    print("\nTesting HUDSON alignment...")
    
    # Create a test workbook to verify alignment
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    
    wb = Workbook()
    ws = wb.active
    
    # Write HUDSON to cell F2
    ws.cell(row=2, column=6, value='HUDSON')
    
    # Apply our formatting function
    ep.apply_master_formats(ws, 2)
    
    # Check if the alignment was applied
    cell_f2 = ws.cell(row=2, column=6)
    
    if cell_f2.alignment and cell_f2.alignment.horizontal == 'right':
        print("✓ HUDSON right-alignment works correctly!")
        return True
    else:
        print(f"✗ HUDSON alignment failed. Alignment: {cell_f2.alignment}")
        return False

if __name__ == "__main__":
    print("Testing Load Factor and HUDSON alignment fixes...\n")
    
    test1_passed = test_load_factor_parsing()
    test2_passed = test_hudson_alignment()
    
    print(f"\nTest Results:")
    print(f"Load Factor parsing: {'PASS' if test1_passed else 'FAIL'}")
    print(f"HUDSON alignment: {'PASS' if test2_passed else 'FAIL'}")
    
    if test1_passed and test2_passed:
        print("\n🎉 All tests passed! The fixes are working correctly.")
    else:
        print("\n❌ Some tests failed. Please check the implementation.")
