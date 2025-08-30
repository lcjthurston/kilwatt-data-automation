import sys
from pathlib import Path
import re

try:
    import pandas as pd
except Exception as e:
    print('DEPENDENCY_ERROR')
    print('pandas is not available:', e)
    sys.exit(10)

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except Exception as e:
    print('DEPENDENCY_ERROR')
    print('openpyxl is not available:', e)
    sys.exit(11)

TARGET_TERMS = {12, 24, 36, 48, 60}

SRC_NAME = '2-copy-reformat/ERCOT-new.xlsx'
DST_NAME = '2-copy-reformat/ERCOT-new-product-term-formulas.xlsx'

BASE_COLS = [
    'Start Month',  # A
    'State',        # B
    'Utility',      # C
    'Congestion Zone',  # D
    'Load Factor',  # E
    'Term',         # F
    'Product',      # G
    '0-200,000',    # H
]


def parse_term_to_int(val):
    if pd.isna(val):
        return None
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        try:
            ival = int(val)
            return ival if float(ival) == val else None
        except Exception:
            return None
    s = str(val).strip()
    if s.isdigit():
        return int(s)
    m = re.search(r"(\d+)", s)
    return int(m.group(1)) if m else None


def find_col(df, preferred_names):
    norm_map = {str(col).strip().lower(): col for col in df.columns}
    for name in preferred_names:
        key = name.strip().lower()
        if key in norm_map:
            return norm_map[key]
    # fallback contains
    for key, orig in norm_map.items():
        if any(name.strip().lower() in key for name in preferred_names):
            return orig
    return None


def filter_sheet(df):
    # Map Product
    prod_col = find_col(df, ['Product', 'Products'])
    term_col = find_col(df, ['Term', 'Terms'])

    if prod_col is None or term_col is None:
        return df.iloc[0:0][BASE_COLS]

    prod_mask = df[prod_col].astype(str).str.strip().str.lower().eq('fixed price')
    term_vals = df[term_col].map(parse_term_to_int)
    term_mask = term_vals.isin(TARGET_TERMS)

    # Select base columns in order, only those that exist
    cols = [c for c in BASE_COLS if c in df.columns]
    filtered = df.loc[prod_mask & term_mask, cols].copy()
    return filtered


def add_formulas(path):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        max_row = ws.max_row
        if max_row < 2:
            continue
        # Optional headers for new columns to help readability
        headers = {
            9: 'I (blank)',
            10: 'J Index',
            11: 'K Concat',
            12: 'L ConstDate',
            13: 'M =B',
            14: 'N Region',
            15: 'O LF Norm',
            16: 'P Supplier',
            17: 'Q TermMonths',
            18: 'R (skip)',
            19: 'S',
            20: 'T',
            21: 'U',
            22: 'V',
            23: 'W',
            24: 'X',
            25: 'Y',
            26: 'Z',
            27: 'AA',
        }
        for col_idx, name in headers.items():
            ws.cell(row=1, column=col_idx, value=name)

        for r in range(2, max_row + 1):
            # J: index (J2 = 1; Jn = J(n-1)+1)
            if r == 2:
                ws.cell(row=r, column=10, value=1)
            else:
                ws.cell(row=r, column=10, value=f"=J{r-1}+1")

            # K: CONCATENATE(C,D)
            ws.cell(row=r, column=11, value=f"=CONCATENATE(C{r},D{r})")

            # L: date constant
            ws.cell(row=r, column=12, value=f"=DATE(2025,8,18)")

            # M: =B
            ws.cell(row=r, column=13, value=f"=A{r}")

            # N: mapping by K
            ws.cell(
                row=r,
                column=14,
                value=(
                    f"=IF(K{r}=\"CenterpointHouston LZ\",\"COAST\"," 
                    f"IF(K{r}=\"OncorNorth LZ\",\"NORTH\"," 
                    f"IF(K{r}=\"AEP TX CENTRALSouth LZ\",\"SOUTH\"," 
                    f"IF(K{r}=\"AEP TX CentralWest LZ\",\"WEST\"," 
                    f"IF(K{r}=\"TNMPHouston LZ\",\"TNMP\",\"NA\")))))"
                ),
            )

            # O: load factor normalization
            ws.cell(row=r, column=15, value=f"=IF(E{r}=\"LO\",\"LOW\",IF(E{r}=\"MED\",\"MED\",IF(E{r}=\"HI\",\"HIGH\",\"NA\")))")

            # P: supplier when fixed price
            ws.cell(row=r, column=16, value=f"=IF(G{r}=\"Fixed Price\",\"APG&E\",\"NA\")")

            # Q: term months integer
            ws.cell(row=r, column=17, value=f"=IF(F{r}=\"12 Months\",12,IF(F{r}=\"24 Months\",24,IF(F{r}=\"36 Months\",36,IF(F{r}=\"48 Months\",48,IF(F{r}=\"60 Months\",60,0)))))")

            # R: skip for now (leave blank)
            # ws.cell(row=r, column=18, value=None)

            # S: constant 200
            ws.cell(row=r, column=19, value="=200")

            # T: =IF(N="",0,L*10)
            ws.cell(row=r, column=20, value=f"=IF(N{r}=\"\",0,H{r}*10)")

            # U: 0
            ws.cell(row=r, column=21, value="=0")

            # V: T+U
            ws.cell(row=r, column=22, value=f"=T{r}+U{r}")

            # W, X, Y, Z: 0
            ws.cell(row=r, column=23, value="=0")
            ws.cell(row=r, column=24, value="=0")
            ws.cell(row=r, column=25, value="=0")
            ws.cell(row=r, column=26, value="=0")

            # AA: 10
            ws.cell(row=r, column=27, value="=10")

    wb.save(path)


def main():
    root = Path('.')
    src = root / SRC_NAME
    if not src.exists():
        print('NOT_FOUND')
        print(f'{SRC_NAME} not found')
        sys.exit(3)

    # Load all sheets and filter
    try:
        sheets = pd.read_excel(src, sheet_name=None)
    except Exception as e:
        print('READ_ERROR')
        print(str(e))
        sys.exit(2)

    if not sheets:
        print('READ_ERROR')
        print('No sheets found')
        sys.exit(2)

    # Write initial A-H
    with pd.ExcelWriter(root / DST_NAME, engine='openpyxl') as writer:
        for name, df in sheets.items():
            filtered = filter_sheet(df)
            # Ensure all base columns exist, even if empty
            for col in BASE_COLS:
                if col not in filtered.columns:
                    filtered[col] = pd.Series(dtype='object')
            filtered = filtered[BASE_COLS]
            filtered.to_excel(writer, sheet_name=str(name)[:31], index=False)

    # Add Excel formulas to columns J..AA (skipping R)
    add_formulas(root / DST_NAME)

    print('SUCCESS')
    print(f'Output written to {DST_NAME}')


if __name__ == '__main__':
    main()

