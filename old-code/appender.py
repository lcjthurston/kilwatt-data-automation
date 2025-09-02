from pathlib import Path
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel as excel_from_serial

# Set source and destination file paths
SRC = Path('2-copy-reformat/ERCOT-filtered.xlsx')  # Path to the source Excel file
DST = Path('2-copy-reformat/DAILY PRICING - new - Copy.xlsx')  # Path to the destination file to append to

# Define column ranges for the source file
SRC_FIRST_COL = 12  # Corresponds to column 'L'
SRC_LAST_COL = 27   # Corresponds to column 'AA'
DST_START_COL = 2   # Start pasting data at column 'B' in the destination file


def last_data_row(ws, scan_cols=40):
    # This function finds the last row containing any data in a worksheet
    # It scans from the bottom up, checking the first 'scan_cols' columns
    max_r = ws.max_row
    for r in range(max_r, 0, -1):
        # Iterate through rows from the last row to the first
        for c in range(1, scan_cols + 1):
            # Iterate through columns to check for data
            v = ws.cell(row=r, column=c).value
            if v not in (None, ''):
                # If a non-empty cell is found, return its row number
                return r
    return 0  # If no data is found, return 0


def gather_source_rows(ws):
    # This function extracts rows from a specific column range in the source worksheet
    rows = []
    # Start scanning from row 2 to skip the header
    for r in range(2, ws.max_row + 1):
        # Get all cell values from the defined source column range (L to AA) for the current row
        row_vals = [ws.cell(row=r, column=c).value for c in range(SRC_FIRST_COL, SRC_LAST_COL + 1)]
        if any(v not in (None, '') for v in row_vals):
            # If any value in the row range is not empty, append the list of values to 'rows'
            rows.append(row_vals)
        else:
            # This 'else' block is a placeholder and doesn't do anything
            # The loop continues to allow for sporadic blank rows, but the outer loop will
            # implicitly stop when the end of the worksheet is reached
            pass
    return rows  # Return the collected rows


def main():
    # Main function to run the script logic
    # Check if both source and destination files exist
    if not SRC.exists():
        print(f"ERROR: Source not found: {SRC}")
        return
    if not DST.exists():
        print(f"ERROR: Destination not found: {DST}")
        return

    # Load the source workbook with data only, ignoring formulas
    wb_src = load_workbook(SRC, data_only=True)
    ws_src = wb_src.worksheets[0]

    # Load the destination workbook (formulas are preserved by default)
    wb_dst = load_workbook(DST)
    ws_dst = wb_dst.worksheets[0]

    # Gather data from the source file
    src_rows = gather_source_rows(ws_src)
    print(f"Source rows to append (L..AA): {len(src_rows)}")

    # Exit if there is no data to append
    if not src_rows:
        print("No data to append.")
        return

    # Find the last data-filled row in the destination file
    last_row = last_data_row(ws_dst, scan_cols=50)
    # Determine the starting row for pasting, which is the row after the last data row
    start_row = last_row + 1 if last_row >= 1 else 1
    print(f"Destination last data row: {last_row}. Appending starting at row {start_row} in columns A..R")

    # Loop through each row from the source and append it to the destination
    for i, row_vals in enumerate(src_rows, start=0):
        # Calculate dest row number
        r = start_row + i
        
        # This section handles the automatic row numbering in column A
        if r == 1:
            # If it's the first row, set the value of cell A1 to 1
            ws_dst.cell(row=r, column=1, value=1)
        else:
            # For subsequent rows, get the value from the previous row's column A
            prev = ws_dst.cell(row=r-1, column=1).value
            try:
                # Attempt to convert the previous cell's value to an integer
                prev_int = int(prev) if prev is not None and str(prev).strip() != '' else 0
            except Exception:
                # If conversion fails (e.g., non-numeric value), default to 0
                prev_int = 0
            # Set the current row's column A value to the previous value + 1
            ws_dst.cell(row=r, column=1, value=prev_int + 1)
        
        # Prepare source row for writing and apply special mapping: swap O and P (indices 3 and 4 within L..AA)
        row_vals_to_write = list(row_vals)
        if len(row_vals_to_write) >= 5:
            row_vals_to_write[3], row_vals_to_write[4] = row_vals_to_write[4], row_vals_to_write[3]

        # Paste source row data to destination ('offset' helps map source to destination columns)
        for offset, val in enumerate(row_vals_to_write):
            # Check if value is likely Excel serial date number
            is_date = False
            if isinstance(val, (int, float)) and 20000 < float(val) < 60000:
                try:
                    # Converting serial number to Python datetime object
                    val = excel_from_serial(float(val))
                    is_date = True
                except Exception:
                    # Keep the original value if conversion fails
                    pass
            # Write value to correct destination cell
            cell = ws_dst.cell(row=r, column=DST_START_COL + offset, value=val)
            if is_date:
                # If the value was a date, set the cell's number format to display it correctly
                cell.number_format = 'mm/dd/yyyy'

    try:
        # Saving changes to destination workbook
        wb_dst.save(DST)
    except PermissionError:
        # Handle the error if the destination file is open
        print("ERROR: Could not save destination file. If it is open in Excel, please close it and re-run.")
        return

    # Print summary 
    print(f"Appended {len(src_rows)} rows (16 columns) from L..AA of source to A..R of destination.")
    # Print sample of last appended row 
    last_appended_row = start_row + len(src_rows) - 1
    sample = [ws_dst.cell(row=last_appended_row, column=c).value for c in range(1, 5)]
    print(f"Last appended row {last_appended_row} sample A..D: {sample}")


# standard python entry point
if __name__ == '__main__':
    main()