import sys
import openpyxl
from copy import copy

MASTER_PATH = 'DAILY PRICING - new.xlsx'
FIRST_APPENDED_ID = 14004586  # from previous append step

try:
    wb = openpyxl.load_workbook(MASTER_PATH)
except PermissionError:
    print('ERROR: Could not open master workbook. Please close it in Excel and retry.')
    sys.exit(2)
except Exception as e:
    print(f'ERROR: Unable to open master workbook: {e}')
    sys.exit(2)

# First visible sheet
ws = None
for name in wb.sheetnames:
    sh = wb[name]
    if getattr(sh, 'sheet_state', 'visible') == 'visible':
        ws = sh
        break
if ws is None:
    ws = wb.active

max_row = ws.max_row

# Find first appended row by ID match in column A
start_row = None
for r in range(2, max_row+1):
    v = ws.cell(row=r, column=1).value
    if v == FIRST_APPENDED_ID:
        start_row = r
        break

if start_row is None:
    print('WARNING: Could not locate the first appended row by ID. Applying styles to all data rows from row 2.')
    start_row = 2

# Capture template styles from row 2, columns A..Q
template_cells = [ws.cell(row=2, column=c) for c in range(1, 17+1)]

styled_cells = 0
for r in range(start_row, max_row+1):
    for c in range(1, 17+1):
        dst = ws.cell(row=r, column=c)
        src = template_cells[c-1]
        # Clone core style components
        dst.font = copy(src.font)
        dst.alignment = copy(src.alignment)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format  # direct assign is fine
        dst.protection = copy(src.protection)
        styled_cells += 1

try:
    wb.save(MASTER_PATH)
except PermissionError:
    print('ERROR: Could not save master workbook. Please close it in Excel and retry.')
    sys.exit(3)

print(f'Applied styles from row 2 to rows {start_row}..{max_row} (A..Q). Total cells styled: {styled_cells}.')