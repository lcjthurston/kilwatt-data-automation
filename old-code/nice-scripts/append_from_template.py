"""
Append rows from a TEMPLATE workbook (IMPORT sheet) into the master workbook,
assigning sequential IDs and enforcing the master sheet's number formats.

- Maps columns by header name (robust against column reordering)
- Normalizes template header variants (e.g., Ruc_Nodal -> RUC_Nodal)
- Auto-generates IDs: max existing ID + 1
- Enforces number formats for appended rows (A..Q) based on master configuration

Usage:
  py append_from_template.py \
    --template "1-ERCOT/DAILY PRICING - APGE - TEMPLATE - 2024'.xlsx" \
    --template-sheet IMPORT \
    --master "DAILY PRICING - new.xlsx"
"""
from __future__ import annotations

import sys
import argparse
from typing import Dict, List

import openpyxl
from openpyxl.utils import get_column_letter

DEFAULT_TEMPLATE_PATH = "1-ERCOT/DAILY PRICING - APGE - TEMPLATE - 2024'.xlsx"
DEFAULT_TEMPLATE_SHEET = "IMPORT"
DEFAULT_MASTER_PATH = "DAILY PRICING - new.xlsx"

# Expected master headers in order (B..Q)
MASTER_HEADERS: List[str] = [
    'Price_Date','Date','Zone','REP1','Load','Term','Min_MWh','Max_MWh',
    'Daily_No_Ruc','RUC_Nodal','Daily','Com_Disc','HOA_Disc','Broker_Fee','Meter_Fee','Max_Meters'
]

# Source header normalization map (case-insensitive)
SOURCE_SYNONYMS: Dict[str, str] = {
    'price_date': 'Price_Date',
    'date': 'Date',
    'zone': 'Zone',
    'rep1': 'REP1',
    'load': 'Load',
    'term': 'Term',
    'min_mwh': 'Min_MWh',
    'max_mwh': 'Max_MWh',
    'daily_no_ruc': 'Daily_No_Ruc',
    'ruc_nodal': 'RUC_Nodal',  # normalize template's Ruc_Nodal
    'daily': 'Daily',
    'com_disc': 'Com_Disc',
    'hoa_disc': 'HOA_Disc',
    'broker_fee': 'Broker_Fee',
    'meter_fee': 'Meter_Fee',
    'max_meters': 'Max_Meters',
}

# Master number formats for A..Q (IDs included)
MASTER_FORMATS: Dict[str, str] = {
    'A': 'General',
    'B': 'mm-dd-yy',
    'C': 'mm-dd-yy',
    'D': 'General',
    'E': 'General',
    'F': 'General',
    'G': '* #,##0;* (#,##0);* -00',
    'H': '* #,##0;* (#,##0);* -00',
    'I': 'General',
    'J': '$#,##0.00;($#,##0.00)',
    'K': '$#,##0.00',
    'L': '$#,##0.00',
    'M': '$* #,##0.00;$* (#,##0.00);$* -00',
    'N': '$* #,##0.00;$* (#,##0.00);$* -00',
    'O': '$#,##0.00',
    'P': '$* #,##0.00;$* (#,##0.00);$* -00',
    'Q': '#,##0',
}


def detect_header_row(ws, max_scan_rows: int = 30) -> int:
    """Find the header row by scanning for expected header names."""
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        norm = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                norm.append(v.strip().lower().replace(' ', '_'))
            else:
                norm.append(None)
        if 'price_date' in norm and ('zone' in norm or 'load' in norm) and ('daily' in norm or 'daily_no_ruc' in norm):
            return r
    return 1


def build_source_mapping(ws, header_row: int) -> Dict[str, int]:
    """Map master header names to source column indices by header text."""
    mapping: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        hv = ws.cell(row=header_row, column=c).value
        if not isinstance(hv, str):
            continue
        key = hv.strip().lower().replace(' ', '_')
        if key in SOURCE_SYNONYMS:
            norm_name = SOURCE_SYNONYMS[key]
            if norm_name in MASTER_HEADERS and norm_name not in mapping:
                mapping[norm_name] = c
    return mapping


def apply_master_formats(ws_dst, row_idx: int) -> None:
    """Apply master number formats for columns A..Q to a single row."""
    for c in range(1, 17 + 1):
        col_letter = get_column_letter(c)
        fmt = MASTER_FORMATS.get(col_letter)
        if fmt:
            ws_dst.cell(row=row_idx, column=c).number_format = fmt


def append_from_template(template_path: str, template_sheet: str, master_path: str) -> None:
    # Load source workbook
    try:
        wb_src = openpyxl.load_workbook(template_path, data_only=True, read_only=False)
    except Exception as e:
        print(f"ERROR: Unable to open template: {e}")
        sys.exit(2)
    if template_sheet not in wb_src.sheetnames:
        print(f"ERROR: Sheet '{template_sheet}' not found in template.")
        sys.exit(2)
    ws_src = wb_src[template_sheet]

    # Discover header row and build mapping
    header_row = detect_header_row(ws_src)
    src_map = build_source_mapping(ws_src, header_row)
    missing = [h for h in MASTER_HEADERS if h not in src_map]
    if missing:
        print('ERROR: Missing expected columns in template:', ', '.join(missing))
        sys.exit(2)

    # Open master workbook
    try:
        wb_dst = openpyxl.load_workbook(master_path)
    except PermissionError:
        print('ERROR: Could not open master workbook for writing. Please close it in Excel and retry.')
        sys.exit(3)
    except Exception as e:
        print(f"ERROR: Unable to open master workbook: {e}")
        sys.exit(3)

    # First visible sheet
    ws_dst = None
    for name in wb_dst.sheetnames:
        sh = wb_dst[name]
        if getattr(sh, 'sheet_state', 'visible') == 'visible':
            ws_dst = sh
            break
    if ws_dst is None:
        ws_dst = wb_dst.active

    # Determine current max ID in column A
    max_id = 0
    for r in range(2, ws_dst.max_row + 1):
        v = ws_dst.cell(row=r, column=1).value
        if isinstance(v, (int, float)):
            try:
                vi = int(v)
            except Exception:
                continue
            if vi > max_id:
                max_id = vi
    next_id = max_id + 1

    # Find the first blank ID row (column A) to start writing, ignoring styled-but-empty rows
    first_blank_row = None
    for r in range(2, ws_dst.max_row + 1):
        v = ws_dst.cell(row=r, column=1).value
        if v in (None, ''):
            first_blank_row = r
            break
    if first_blank_row is None:
        first_blank_row = ws_dst.max_row + 1

    # Iterate source rows and append
    start_data_row = header_row + 1
    rows_appended = 0
    preview = []
    write_row = first_blank_row

    for r in range(start_data_row, ws_src.max_row + 1):
        values: List[object] = []
        all_empty = True
        for h in MASTER_HEADERS:
            c_idx = src_map[h]
            cell = ws_src.cell(row=r, column=c_idx)
            v = cell.value
            if isinstance(v, str):
                v = v.strip()
            if v not in (None, ''):
                all_empty = False
            values.append(v)
        if all_empty:
            continue

        dst_row = write_row
        # ID
        ws_dst.cell(row=dst_row, column=1, value=next_id)
        # Apply formats to whole row A..Q first
        apply_master_formats(ws_dst, dst_row)
        # Write values for B..Q in the master-defined order
        for i, val in enumerate(values):
            ws_dst.cell(row=dst_row, column=i + 2, value=val)

        write_row += 1

        if rows_appended < 3:
            preview.append([ws_dst.cell(row=dst_row, column=c).value for c in range(1, 8)])

        rows_appended += 1
        next_id += 1

    if rows_appended == 0:
        print('No non-empty rows found to append.')
        return

    # Save master
    try:
        wb_dst.save(master_path)
    except PermissionError:
        print('ERROR: Could not save master workbook. Please close it in Excel and retry.')
        sys.exit(4)

    print(f'Appended {rows_appended} rows. New ID range: {max_id + 1}..{next_id - 1}')
    print('Preview of first 3 appended rows (A..G):')
    for row in preview:
        print(row)


def main() -> None:
    parser = argparse.ArgumentParser(description='Append rows from a template workbook into the master workbook.')
    parser.add_argument('--template', default=DEFAULT_TEMPLATE_PATH, help='Path to the template .xlsx file')
    parser.add_argument('--template-sheet', default=DEFAULT_TEMPLATE_SHEET, help='Worksheet name in the template')
    parser.add_argument('--master', default=DEFAULT_MASTER_PATH, help='Path to the master .xlsx file')
    args = parser.parse_args()

    append_from_template(args.template, args.template_sheet, args.master)


if __name__ == '__main__':
    main()

