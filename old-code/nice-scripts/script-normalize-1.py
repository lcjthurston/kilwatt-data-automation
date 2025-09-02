# Normalize the Term column to integers
import re, sys
import openpyxl
from openpyxl.utils import get_column_letter

MASTER_PATH = 'DAILY PRICING - new.xlsx'
TERM_COL_IDX = 7  # Column G
TERM_FORMAT = '* #,##0;* (#,##0);* -00'

try:
    wb = openpyxl.load_workbook(MASTER_PATH)
except PermissionError:
    print('ERROR: Could not open master workbook. Please close it in Excel and retry.')
    sys.exit(2)
except Exception as e:
    print(f'ERROR: Unable to open master workbook: {e}')
    sys.exit(2)

# First visible sheet
ws = None
for name in wb.sheetnames:
    sh = wb[name]
    if getattr(sh, 'sheet_state', 'visible') == 'visible':
        ws = sh
        break
if ws is None:
    ws = wb.active

# Detect header row (assume row 1 or first row where 'Term' appears in column G)
header_row = 1
header_val = ws.cell(row=header_row, column=TERM_COL_IDX).value
if not (isinstance(header_val, str) and header_val.strip().lower() == 'term'):
    # search first 10 rows
    for r in range(1, min(ws.max_row, 10)+1):
        v = ws.cell(row=r, column=TERM_COL_IDX).value
        if isinstance(v, str) and v.strip().lower() == 'term':
            header_row = r
            break

updated = 0
converted_samples = []
pattern = re.compile(r'(\d+)')

for r in range(header_row+1, ws.max_row+1):
    cell = ws.cell(row=r, column=TERM_COL_IDX)
    val = cell.value
    # Ensure number format regardless
    cell.number_format = TERM_FORMAT

    if isinstance(val, str):
        m = pattern.search(val)
        if m:
            n = int(m.group(1))
            if val != n:
                cell.value = n
                updated += 1
                if len(converted_samples) < 5:
                    converted_samples.append((r, val, n))

try:
    wb.save(MASTER_PATH)
except PermissionError:
    print('ERROR: Could not save master workbook. Please close it in Excel and retry.')
    sys.exit(3)

print(f'Updated Term values: {updated} cells converted to integers.')
if converted_samples:
    print('Sample conversions:')
    for row, before, after in converted_samples:
        print(f'  Row {row}: {before!r} -> {after}')
