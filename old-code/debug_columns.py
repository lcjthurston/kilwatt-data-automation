#!/usr/bin/env python3

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

def examine_master_table():
    """Examine the master table structure to understand column layout"""
    
    master_path = Path('2-copy-reformat/DAILY PRICING - master table.xlsx')
    
    if not master_path.exists():
        print(f"Master table not found: {master_path}")
        return
    
    # Read with pandas to see the column structure
    print("=== PANDAS READ ===")
    try:
        df = pd.read_excel(master_path)
        print(f"Columns found by pandas: {list(df.columns)}")
        print(f"Number of columns: {len(df.columns)}")
        print(f"First few rows:")
        print(df.head())
    except Exception as e:
        print(f"Error reading with pandas: {e}")
    
    print("\n=== OPENPYXL READ ===")
    # Read with openpyxl to see the raw cell structure
    try:
        wb = load_workbook(master_path)
        ws = wb.active
        
        print(f"Active sheet name: {ws.title}")
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
        
        # Check the first row (headers)
        print("\nFirst row (headers):")
        for col in range(1, min(20, ws.max_column + 1)):  # Check first 20 columns
            cell_value = ws.cell(row=1, column=col).value
            print(f"  Column {col} ({chr(64+col)}): '{cell_value}'")
        
        # Check a few data rows
        print(f"\nFirst few data rows:")
        for row in range(2, min(5, ws.max_row + 1)):
            print(f"Row {row}:")
            for col in range(1, min(10, ws.max_column + 1)):  # Check first 10 columns
                cell_value = ws.cell(row=row, column=col).value
                print(f"  Column {col} ({chr(64+col)}): '{cell_value}'")
        
        wb.close()
        
    except Exception as e:
        print(f"Error reading with openpyxl: {e}")

if __name__ == "__main__":
    examine_master_table()
