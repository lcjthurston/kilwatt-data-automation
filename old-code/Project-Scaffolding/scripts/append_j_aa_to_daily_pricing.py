from pathlib import Path
from datetime import date
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel as excel_date

# Source and destination within the 2-copy-reformat folder
SRC = Path('2-copy-reformat/ERCOT-filtered.xlsx')
DST = Path('2-copy-reformat/DAILY PRICING - new - Copy.xlsx')

# Columns A-H source meanings
# A: Start Month, B: State, C: Utility, D: Congestion Zone, E: Load Factor, F: Term, G: Product, H: 0-200,000


def parse_term_to_int(val):
    if val is None:
        return None
    if isinstance(val, (int,)):
        return val
    if isinstance(val, float):
        iv = int(val)
        return iv if float(iv) == val else None
    s = str(val).strip()
    m = re.search(r"(\d+)", s)
    return int(m.group(1)) if m else None


def norm_load_factor(v):
    if v is None:
        return 'NA'
    s = str(v).strip()
    # try direct codes first
    s_up = s.upper()
    if s_up == 'LO':
        return 'LOW'
    if s_up == 'MED':
        return 'MED'
    if s_up == 'HI':
        return 'HIGH'
    # attempt to extract code from strings like 'RESIDENTIAL LO'
    m = re.search(r"(LO|MED|HI)(?:\b|$)", s_up)
    if not m:
        return 'NA'
    code = m.group(1)
    return {'LO': 'LOW', 'MED': 'MED', 'HI': 'HIGH'}.get(code, 'NA')


def region_from_concat(k):
    if k == 'CenterpointHouston LZ':
        return 'COAST'
    if k == 'OncorNorth LZ':
        return 'NORTH'
    if k == 'AEP TX CENTRALSouth LZ':
        return 'SOUTH'
    if k == 'AEP TX CentralWest LZ':
        return 'WEST'
    if k == 'TNMPHouston LZ':
        return 'TNMP'
    return 'NA'


def compute_values_from_row(a,b,c,d,e,f,g,h, row_index):
    # J..AA values according to spec (R left blank)
    J = row_index  # start at 1 for first data row (row 2 in Excel)
    K = f"{c or ''}{d or ''}"
    L = date(2025,8,18)
    M = b
    N = region_from_concat(K)
    O = norm_load_factor(e)
    P = 'APG&E' if str(g).strip().lower() == 'fixed price' else 'NA'
    term_int = parse_term_to_int(f)
    Q = term_int if term_int in {12,24,36,48,60} else 0
    R = None  # skipped
    S = 200
    # T uses L*10 when N not blank
    if N:
        T = excel_date(L) * 10
    else:
        T = 0
    U = 0
    V = T + U
    W = 0
    X = 0
    Y = 0
    Z = 0
    AA = 10
    return [J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z,AA]


def main():
    if not SRC.exists():
        print(f"ERROR: Source not found: {SRC}")
        return
    if not DST.exists():
        print(f"ERROR: Target not found: {DST}")
        return

    # Load source with data_only=True to fetch cached values if present
    wb_src_vals = load_workbook(SRC, data_only=True)
    ws_src_vals = wb_src_vals.worksheets[0]

    # Also load formulas workbook to read base columns if we need to compute fallbacks
    wb_src_form = load_workbook(SRC, data_only=False)
    ws_src_form = wb_src_form.worksheets[0]

    # Load target workbook for writing
    wb_dst = load_workbook(DST)
    ws_dst = wb_dst.worksheets[0]

    # Build lists of data row indices in source and destination (based on any value in A..H)
    src_rows = []
    for r in range(2, ws_src_form.max_row + 1):
        if any(ws_src_form.cell(row=r, column=col).value not in (None, '') for col in range(1, 9)):
            src_rows.append(r)
        else:
            break

    dst_rows = []
    for r in range(2, ws_dst.max_row + 1):
        if any(ws_dst.cell(row=r, column=col).value not in (None, '') for col in range(1, 9)):
            dst_rows.append(r)
        else:
            break

    rows_to_write = min(len(src_rows), len(dst_rows))
    print(f"Rows to copy (bounded by target): {rows_to_write} (source has {len(src_rows)}, target has {len(dst_rows)})")

    # Loop over only the number of rows present in the destination
    for idx, (r_src, r_dst) in enumerate(zip(src_rows[:rows_to_write], dst_rows[:rows_to_write]), start=1):
        # Try to read precomputed values J..AA from source row r_src
        vals = []
        all_none = True
        for c in range(10, 28):
            v = ws_src_vals.cell(row=r_src, column=c).value
            vals.append(v)
            if v not in (None, ''):
                all_none = False
        if all_none:
            # Compute ourselves from A..H of the source row r_src
            a = ws_src_form.cell(row=r_src, column=1).value
            b = ws_src_form.cell(row=r_src, column=2).value
            c_val = ws_src_form.cell(row=r_src, column=3).value
            d = ws_src_form.cell(row=r_src, column=4).value
            e = ws_src_form.cell(row=r_src, column=5).value
            f = ws_src_form.cell(row=r_src, column=6).value
            g = ws_src_form.cell(row=r_src, column=7).value
            h = ws_src_form.cell(row=r_src, column=8).value
            vals = compute_values_from_row(a,b,c_val,d,e,f,g,h, idx)
        # Write values to destination row r_dst
        for offset, c in enumerate(range(10,28)):
            ws_dst.cell(row=r_dst, column=c).value = vals[offset]

    wb_dst.save(DST)
    print(f"Appended J..AA values into '{DST.name}' on first sheet.")

if __name__ == '__main__':
    main()

