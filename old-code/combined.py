import sys
from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Define constants
TARGET_TERMS = {12, 24, 36, 48, 60}
SRC_NAME = '2-copy-reformat/ERCOT-unfiltered-source.xlsx' # Assuming this is your unfiltered source
DST_NAME = '2-copy-reformat/DAILY PRICING - master table.xlsx' # This is your master table

BASE_COLS = [
    'Start Month',
    'State',
    'Utility',
    'Congestion Zone',
    'Load Factor',
    'Term',
    'Product',
    '0-200,000',
]

def parse_term_to_int(val):
    if pd.isna(val):
        return None
    if isinstance(val, int):
        return val
    s = str(val).strip()
    if s.isdigit():
        return int(s)
    return None

def filter_sheet(df):
    df_filtered = pd.DataFrame(columns=BASE_COLS)
    for col in BASE_COLS:
        if col in df.columns:
            df_filtered[col] = df[col]
        else:
            df_filtered[col] = pd.NA
    
    # Apply the filtering criteria from the original script
    df_filtered = df_filtered[
        df_filtered['Utility'].str.upper().isin(['CPL', 'AEPN', 'ONCOR', 'TNMP']) &
        (df_filtered['Load Factor'] == '0-100%') &
        (df_filtered['Term'].apply(parse_term_to_int).isin(TARGET_TERMS))
    ]
    return df_filtered

def last_data_row(ws, scan_cols=40):
    max_r = ws.max_row
    for r in range(max_r, 0, -1):
        for c in range(1, scan_cols + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ''):
                return r
    return 0

def main():
    root = Path('.')
    src = root / SRC_NAME
    dst = root / DST_NAME

    if not src.exists():
        print(f"ERROR: Source not found: {src}")
        sys.exit(3)
    if not dst.exists():
        print(f"ERROR: Destination not found: {dst}")
        sys.exit(3)

    # Load source data
    try:
        sheets = pd.read_excel(src, sheet_name=None)
    except Exception as e:
        print('READ_ERROR')
        print(str(e))
        sys.exit(2)

    # Load destination workbook to prepare for appending
    wb_dst = load_workbook(dst)
    ws_dst = wb_dst.active
    
    last_row = last_data_row(ws_dst)
    start_row = last_row + 1 if last_row >= 1 else 1

    # Capture the formats from the last existing row of the destination file
    dst_formats = {}
    if last_row > 0:
        for c in range(1, len(BASE_COLS) + 1):
            dst_formats[c] = ws_dst.cell(row=last_row, column=c).number_format
    
    # Process and append data
    for name, df in sheets.items():
        filtered = filter_sheet(df)
        
        for r_offset, row_data in enumerate(filtered.itertuples(index=False), start=0):
            r = start_row + r_offset
            
            # Append values to the destination worksheet
            for c_offset, value in enumerate(row_data, start=1):
                cell = ws_dst.cell(row=r, column=c_offset, value=value)
                # Apply the captured format, defaulting to 'General' if not found
                cell.number_format = dst_formats.get(c_offset, 'General')

    # Add formulas to the appended rows
    # NOTE: This part needs to be customized to your specific formula logic from your second script
    # This is an example, assuming formulas are in columns 9 through 27
    for r in range(start_row, ws_dst.max_row + 1):
        # Example formula logic for columns I through AA
        ws_dst.cell(row=r, column=9, value=f"=O{r}+P{r}") # Example for I
        ws_dst.cell(row=r, column=10, value=f"=Q{r}*R{r}") # Example for J
        # ... and so on for all your formula columns ...

    try:
        wb_dst.save(dst)
        print(f"Successfully appended {ws_dst.max_row - last_row} rows to {DST_NAME}.")
    except PermissionError:
        print("ERROR: Could not save destination file. Please close it and re-run.")
        sys.exit(4)

if __name__ == '__main__':
    main()